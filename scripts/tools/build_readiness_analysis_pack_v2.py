from __future__ import annotations

import hashlib
import re
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from scripts.parsers.learning_parser import _load_learning_database, _load_roi_catalog
from scripts.rtm_utils import load_login_employee_map
from scripts.utils import load_settings


ROOT = Path(__file__).resolve().parents[2]
OUT = ROOT / "data" / "out"
RAW = ROOT / "data" / "raw"
REPORTS = ROOT / "reports"
OUTPUT = REPORTS / "readiness_product_analysis.xlsx"
COURSE_CATALOG = ROOT / "config" / "courses_catalog_ЛМ_ROI_пример.xlsx"

PROJECT_ID = "1054"
PROJECT_NAME = "FMCG-проект"
KPI_TARGET = 0.90
OKK_TARGET = 0.50
FRAUD_MAX = 0.20
TRAINING_TARGET = 0.90
TEST_DEFAULT_TARGET = 0.90
ANALYSIS_START = pd.Timestamp("2026-01-01")
ANALYSIS_END = pd.Timestamp("2026-07-31")
COHORT_START = pd.Timestamp("2026-01-01")
COHORT_END = pd.Timestamp("2026-06-30")
MAY_START = pd.Timestamp("2026-05-01")
MAY_END = pd.Timestamp("2026-05-31")
JUNE_END = pd.Timestamp("2026-06-30")
JULY_END = pd.Timestamp("2026-07-31")
RTM_START = ANALYSIS_START
RTM_END = JULY_END
RTM_REQUIRED_MONTHS = tuple(202600 + month for month in range(1, 8))


def first_notna(series: pd.Series):
    clean = series.dropna()
    return clean.iloc[0] if not clean.empty else pd.NA


def mode_or_first(series: pd.Series):
    clean = series.dropna().astype("string").str.strip()
    clean = clean[clean.ne("")]
    if clean.empty:
        return pd.NA
    mode = clean.mode()
    return mode.iloc[0] if not mode.empty else clean.iloc[0]


def num(series: pd.Series) -> pd.Series:
    return pd.to_numeric(series, errors="coerce")


def date(series: pd.Series) -> pd.Series:
    return pd.to_datetime(series, errors="coerce").dt.normalize()


def anonymizer(values: pd.Series, prefix: str, width: int) -> dict[str, str]:
    unique = sorted({str(value).strip() for value in values.dropna() if str(value).strip()})
    hash_length = max(width, 10)
    mapping = {
        value: f"{prefix}_{hashlib.sha256(f'HN|{prefix}|{value}'.encode('utf-8')).hexdigest()[:hash_length].upper()}"
        for value in unique
    }
    if len(set(mapping.values())) != len(mapping):
        raise ValueError(f"Collision in anonymized {prefix} identifiers")
    return mapping


def anon(value, mapping: dict[str, str], missing: str) -> str:
    if pd.isna(value) or not str(value).strip():
        return missing
    return mapping.get(str(value).strip(), missing)


def episode_key(employee_id: str, project_id: str, hire_date) -> str:
    hire = pd.to_datetime(hire_date, errors="coerce")
    hire_label = hire.strftime("%Y%m%d") if pd.notna(hire) else "NO_DATE"
    return f"{employee_id}|{project_id}|{hire_label}"


def stable_hash(*values, prefix="EVT") -> str:
    raw = "|".join("" if pd.isna(value) else str(value) for value in values)
    return f"{prefix}_{hashlib.sha1(raw.encode('utf-8')).hexdigest()[:14].upper()}"


def status_at(row: pd.Series, checkpoint: pd.Timestamp) -> str:
    start = row["Дата приёма на проект"]
    end = row["Дата увольнения с проекта"]
    if pd.isna(start) or start > checkpoint:
        return "Не вышел"
    if pd.notna(end) and end <= checkpoint:
        confirmed_exit = row.get("Увольнение подтверждено", False)
        if pd.notna(confirmed_exit) and bool(confirmed_exit):
            return "Уволен"
        return "Конфликт кадровых данных"
    return "Активен на проекте"


def load_facts():
    kpi_path = OUT / "kpi_employee_history_fact.parquet"
    if not kpi_path.exists():
        kpi_path = OUT / "kpi_fact.parquet"
    facts = {
        "employees": pd.read_parquet(OUT / "dim_employees.parquet"),
        "hr": pd.read_parquet(OUT / "fact_hr_registry.parquet"),
        "kpi": pd.read_parquet(kpi_path),
        "kpi_tt": pd.read_parquet(OUT / "kpi_client_tt_fact.parquet"),
        "kpi_long": pd.read_parquet(OUT / "kpi_client_tt_long.parquet"),
        "okk": pd.read_parquet(OUT / "okk_fact.parquet"),
    }
    facts["hr"]["Дата приема"] = date(facts["hr"]["Дата приема"])
    facts["hr"]["Дата увольнения"] = date(facts["hr"]["Дата увольнения"])
    facts["kpi"]["MonthStart"] = date(facts["kpi"]["MonthStart"])
    facts["kpi_tt"]["MonthStart"] = date(facts["kpi_tt"]["MonthStart"])
    facts["kpi_long"]["MonthStart"] = date(facts["kpi_long"]["MonthStart"])
    facts["okk"]["Дата визита"] = date(facts["okk"]["Дата визита"])
    for column in ["KPI 1", "KPI 2", "Сервис (факт)", "ОСА (факт)", "PICoS (факт)", "Покрытие (факт)"]:
        if column in facts["kpi"]:
            facts["kpi"][column] = num(facts["kpi"][column])
    for column in facts["okk"].columns:
        if column in {
            "Качество визита",
            "Качество 1-ого фотоаудита",
            "% качества PICoS",
            "% наличия товара на полке",
            "% наличия PICoS",
        }:
            facts["okk"][column] = num(facts["okk"][column])
    facts["okk"]["Флаг фальсификации"] = facts["okk"]["Флаг фальсификации"].fillna(False).astype(bool)
    return facts


def load_raw_learning() -> pd.DataFrame:
    roi = _load_roi_catalog(ROOT / "config")
    raw = _load_learning_database(load_settings(), roi, active_only=False)
    raw["start_date"] = pd.to_datetime(raw["start_date"], errors="coerce")
    raw["completion_date"] = pd.to_datetime(raw["completion_date"], errors="coerce")
    raw["completion_pct"] = num(raw["completion_pct"])
    raw["test_score"] = num(raw["test_score"])
    raw = raw.sort_values(
        ["employee_id_raw", "course_id", "completion_pct", "completion_date"],
        ascending=[True, True, False, False],
        na_position="last",
    ).drop_duplicates(["employee_id_raw", "course_id"], keep="first")

    catalog = pd.read_excel(COURSE_CATALOG, sheet_name="Лист1")
    catalog = catalog.rename(
        columns={
            "Номер курса в КУ": "course_id",
            "Название курса в КУ": "course_name_catalog",
            "Обязательный курс": "mandatory_raw",
            "Учавствует в адаптации": "adaptation_raw",
            "Есть тестирование или считать по прогрессу (посещению)": "completion_method",
            "Балл тестирования, который считается успешной сдачей теста": "test_threshold",
        }
    )
    catalog["course_id"] = catalog["course_id"].astype("string").str.strip()
    catalog["test_threshold"] = num(catalog["test_threshold"])
    raw = raw.merge(
        catalog[
            [
                "course_id",
                "course_name_catalog",
                "mandatory_raw",
                "adaptation_raw",
                "completion_method",
                "test_threshold",
            ]
        ],
        on="course_id",
        how="left",
    )
    raw["Обязательный"] = raw["mandatory_raw"].astype("string").str.lower().isin(["да", "true", "1"])
    raw["Входит в адаптацию"] = raw["adaptation_raw"].astype("string").str.lower().isin(["да", "true", "1"])
    method = raw["completion_method"].astype("string").str.lower().fillna("")
    threshold = raw["test_threshold"].fillna(TEST_DEFAULT_TARGET)
    by_test = method.str.contains("тест", na=False) & raw["test_score"].ge(threshold)
    by_progress = method.str.contains("прогресс", na=False) & raw["completion_pct"].ge(1)
    by_attendance = method.str.contains("посещ|был", na=False) & raw["status"].eq("passed")
    fallback = raw["status"].eq("passed") & ~(method.str.contains("тест|прогресс|посещ|был", na=False))
    raw["Успешно завершён"] = (by_test | by_progress | by_attendance | fallback) & raw["completion_date"].notna()
    raw["Название курса"] = raw["course_name_catalog"].combine_first(raw["course_name_src"])
    return raw


def build_maps(facts, raw_learning):
    employee_values = pd.concat(
        [
            facts["employees"]["ID сотрудника"],
            facts["hr"]["ID сотрудника"],
            facts["kpi"]["ID мерчендайзера"],
            facts["okk"]["ID мерчендайзера"],
            raw_learning["employee_id_raw"],
        ],
        ignore_index=True,
    )
    supervisor_values = pd.concat(
        [facts["hr"]["ID супервайзера"], facts["kpi"]["ID супервайзера"], facts["okk"]["ID супервайзера"]],
        ignore_index=True,
    )
    tm_values = pd.concat([facts["hr"]["ID территориального менеджера"], facts["okk"]["ID ТМ"]], ignore_index=True)
    tt_values = pd.concat([facts["okk"]["Код ТТ"], facts["kpi_tt"]["ТТ"]], ignore_index=True)
    return {
        "employee": anonymizer(employee_values, "EMP", 5),
        "supervisor": anonymizer(supervisor_values, "SV", 3),
        "tm": anonymizer(tm_values, "TM", 3),
        "tt": anonymizer(tt_values, "TT", 6),
    }


def build_episodes(hr: pd.DataFrame, maps) -> pd.DataFrame:
    work = hr[hr["ID сотрудника"].notna()].copy()
    work["ID сотрудника raw"] = work["ID сотрудника"].astype("string").str.strip()
    work["ID сотрудника"] = work["ID сотрудника raw"].map(maps["employee"])
    work["ID проекта"] = work["Проект"].astype("string").fillna(PROJECT_ID)
    work["Название проекта"] = PROJECT_NAME
    work["Дата приёма на проект"] = work["Дата приема"]
    work["Дата увольнения с проекта"] = work["Дата увольнения"]
    work["Супервайзер"] = work["ID супервайзера"].map(
        lambda value: anon(value, maps["supervisor"], "SV_UNKNOWN")
    )
    work["Территориальный менеджер"] = work["ID территориального менеджера"].map(
        lambda value: anon(value, maps["tm"], "TM_UNKNOWN")
    )
    work["Дата начала кадровой записи"] = work["Дата приема"]
    work["Дата окончания кадровой записи"] = work["Дата увольнения"]

    group_columns = ["ID сотрудника raw", "ID проекта", "Дата приёма на проект"]
    work["Строк источника в эпизоде"] = work.groupby(group_columns, dropna=False)["ID сотрудника raw"].transform("size")
    work = (
        work.sort_values(["ID сотрудника raw", "ID проекта", "Дата приёма на проект", "Дата увольнения с проекта"])
        .groupby(group_columns, dropna=False)
        .agg(
            **{
                "ID сотрудника": ("ID сотрудника", "first"),
                "Название проекта": ("Название проекта", "first"),
                "Дата увольнения с проекта": ("Дата увольнения с проекта", "max"),
                "Должность": ("Должность", mode_or_first),
                "Регион": ("Регион BI", mode_or_first),
                "Супервайзер": ("Супервайзер", mode_or_first),
                "Территориальный менеджер": ("Территориальный менеджер", mode_or_first),
                "Дата начала кадровой записи": ("Дата начала кадровой записи", "min"),
                "Дата окончания кадровой записи": ("Дата окончания кадровой записи", "max"),
                "Строк источника в эпизоде": ("Строк источника в эпизоде", "max"),
                "Кадровое состояние": ("Состояние", mode_or_first),
                "Активен в USERS": ("Активен в USERS", first_notna),
            }
        )
        .reset_index()
    )
    work = work.sort_values(["ID сотрудника raw", "ID проекта", "Дата приёма на проект"]).reset_index(drop=True)

    types = []
    for _, group in work.groupby(["ID сотрудника raw", "ID проекта"], sort=False):
        previous = None
        for _, row in group.iterrows():
            if previous is None:
                types.append((row.name, "Первичный найм"))
            elif pd.notna(previous["Дата увольнения с проекта"]) and previous["Дата увольнения с проекта"] < row["Дата приёма на проект"]:
                types.append((row.name, "Повторный найм"))
            else:
                types.append((row.name, "Перевод / изменение записи"))
            previous = row
    type_map = dict(types)
    work["Тип выхода"] = work.index.map(type_map)
    work["Дата следующего эпизода"] = work.groupby(["ID сотрудника raw", "ID проекта"])[
        "Дата приёма на проект"
    ].shift(-1)
    state_exit = work["Кадровое состояние"].astype("string").str.contains(
        "увольнение", case=False, na=False
    )
    current_inactive = work["Активен в USERS"].eq(False)
    rehired = (
        work["Дата следующего эпизода"].notna()
        & work["Дата увольнения с проекта"].notna()
        & work["Дата следующего эпизода"].gt(work["Дата увольнения с проекта"])
    )
    work["Увольнение подтверждено"] = (
        work["Дата увольнения с проекта"].notna() & state_exit & (current_inactive | rehired)
    )
    work["Конфликт кадровой даты"] = (
        work["Дата увольнения с проекта"].notna()
        & work["Дата увольнения с проекта"].le(JULY_END)
        & ~work["Увольнение подтверждено"]
    )
    work["Ключ эпизода"] = work.apply(
        lambda row: episode_key(row["ID сотрудника"], row["ID проекта"], row["Дата приёма на проект"]),
        axis=1,
    )
    for label, checkpoint in [
        ("Статус на 31 мая", MAY_END),
        ("Статус на 30 июня", JUNE_END),
        ("Статус на 31 июля", JULY_END),
    ]:
        work[label] = work.apply(lambda row: status_at(row, checkpoint), axis=1)
    work["Причина увольнения"] = pd.NA
    work["Метод дат кадровой записи"] = "Дата начала = дата приёма; дата окончания = дата увольнения из кадрового реестра"
    return work


def assign_episode(employee_raw, event_date, episodes: pd.DataFrame, prefer_future=False):
    candidates = episodes[episodes["ID сотрудника raw"].eq(str(employee_raw).strip())].copy()
    if candidates.empty or pd.isna(event_date):
        return pd.NA
    event_date = pd.Timestamp(event_date).normalize()
    active = candidates[
        candidates["Дата приёма на проект"].le(event_date)
        & (candidates["Дата увольнения с проекта"].isna() | candidates["Дата увольнения с проекта"].ge(event_date))
    ]
    if not active.empty:
        return active.sort_values("Дата приёма на проект").iloc[-1]["Ключ эпизода"]
    if prefer_future:
        future = candidates[candidates["Дата приёма на проект"].gt(event_date)]
        if not future.empty:
            return future.sort_values("Дата приёма на проект").iloc[0]["Ключ эпизода"]
    past = candidates[candidates["Дата приёма на проект"].le(event_date)]
    if not past.empty:
        return past.sort_values("Дата приёма на проект").iloc[-1]["Ключ эпизода"]
    return candidates.sort_values("Дата приёма на проект").iloc[0]["Ключ эпизода"]


def build_learning(raw_learning: pd.DataFrame, episodes: pd.DataFrame, maps) -> pd.DataFrame:
    learning = raw_learning.copy()
    learning["ID сотрудника raw"] = learning["employee_id_raw"].astype("string").str.strip()
    project_employee_ids = set(episodes["ID сотрудника raw"].dropna().astype("string").str.strip())
    learning = learning[learning["ID сотрудника raw"].isin(project_employee_ids)].copy()
    learning["ID сотрудника"] = learning["ID сотрудника raw"].map(maps["employee"]).fillna("EMP_UNKNOWN")
    learning["Проект"] = PROJECT_NAME
    learning["Номер курса"] = learning["course_id"]
    learning["Дата назначения"] = learning["start_date"]
    learning["Дата начала"] = learning["start_date"]
    learning["Дата завершения"] = learning["completion_date"]
    learning["Прогресс"] = learning["completion_pct"]
    learning["Балл теста"] = learning["test_score"]
    learning["Ключ эпизода"] = learning.apply(
        lambda row: assign_episode(row["ID сотрудника raw"], row["Дата назначения"], episodes, prefer_future=True),
        axis=1,
    )
    hire_lookup = episodes.set_index("Ключ эпизода")["Дата приёма на проект"].to_dict()
    learning["Дата приёма связанного эпизода"] = learning["Ключ эпизода"].map(hire_lookup)
    learning["Предобучение"] = learning["Дата назначения"].dt.normalize().lt(learning["Дата приёма связанного эпизода"])
    learning["Дней относительно выхода"] = (
        learning["Дата назначения"].dt.normalize() - learning["Дата приёма связанного эпизода"]
    ).dt.days
    learning["Метод даты назначения"] = "Отдельной даты назначения нет; используется дата начала обучения"
    columns = [
        "ID сотрудника",
        "Ключ эпизода",
        "Проект",
        "Номер курса",
        "Название курса",
        "Обязательный",
        "Входит в адаптацию",
        "Дата назначения",
        "Дата начала",
        "Дата завершения",
        "Прогресс",
        "Балл теста",
        "Успешно завершён",
        "Предобучение",
        "Дней относительно выхода",
        "Метод даты назначения",
    ]
    return learning[columns].sort_values(["ID сотрудника", "Дата назначения", "Номер курса"])


def formation_dates() -> dict[int, pd.Timestamp]:
    result = {}
    for path in (RAW / "kpi" / "fact kpi").glob("*.xlsx"):
        period_match = re.search(r"(?<!\d)(0?[1-9]|1[0-2])\.2026(?!\d)", path.name)
        date_match = re.search(r"от\s*(\d{1,2})\.(\d{1,2})\.(\d{2,4})", path.name, flags=re.IGNORECASE)
        if not period_match:
            continue
        month = int(period_match.group(1))
        if date_match:
            year = int(date_match.group(3))
            year = 2000 + year if year < 100 else year
            result[202600 + month] = pd.Timestamp(year, int(date_match.group(2)), int(date_match.group(1)))
        else:
            result[202600 + month] = pd.Timestamp.fromtimestamp(path.stat().st_mtime).normalize()
    return result


def build_kpi_monthly(facts, episodes: pd.DataFrame, maps) -> pd.DataFrame:
    kpi = facts["kpi"].copy()
    kpi = kpi[kpi["YearMonth"].between(202601, 202607)].copy()
    if "Визиты KPI" not in kpi.columns:
        kpi["Визиты KPI"] = pd.NA
    if "ТТ KPI" not in kpi.columns:
        kpi["ТТ KPI"] = pd.NA
    grouped = (
        kpi.groupby(["ID мерчендайзера", "MonthStart", "YearMonth"], dropna=False)
        .agg(
            **{
                "Итоговый KPI": ("KPI 1", "mean"),
                "KPI 2": ("KPI 2", "mean"),
                "Сервис факт": ("Сервис (факт)", "mean"),
                "OSA факт": ("ОСА (факт)", "mean"),
                "PICOS факт": ("PICoS (факт)", "mean"),
                "Покрытие факт": ("Покрытие (факт)", "mean"),
                "Покрытие план": ("Покрытие (план)", "mean"),
                "PICOS СВ факт": ("PICoS СВ (факт)", "mean"),
                "PICOS СВ план": ("PICoS СВ (план)", "mean"),
                "Количество подтверждённых визитов RTM": ("Визиты KPI", "sum"),
                "Количество ТТ с клиентским KPI": ("ТТ KPI", "sum"),
                "Регион": ("Регион BI", mode_or_first),
                "ID супервайзера raw": ("ID супервайзера", first_notna),
            }
        )
        .reset_index()
    )
    grouped["ID сотрудника raw"] = grouped["ID мерчендайзера"].astype("string").str.strip()
    grouped["ID сотрудника"] = grouped["ID сотрудника raw"].map(maps["employee"]).fillna("EMP_UNKNOWN")
    grouped["Проект"] = PROJECT_NAME
    grouped["Месяц"] = grouped["MonthStart"]
    grouped["Рабочий порог KPI"] = KPI_TARGET
    grouped["Окончательный расчёт месяца"] = True
    grouped["Дата формирования показателя"] = grouped["YearMonth"].map(formation_dates())
    grouped["Руководитель в месяце"] = grouped["ID супервайзера raw"].map(
        lambda value: anon(value, maps["supervisor"], "SV_UNKNOWN")
    )
    grouped["Ключ эпизода"] = grouped.apply(
        lambda row: assign_episode(
            row["ID сотрудника raw"], row["MonthStart"] + pd.offsets.MonthEnd(0), episodes
        ),
        axis=1,
    )
    grouped["Метод распределения"] = "Клиентский KPI ТТ связан с подтверждёнными RTM-визитами сотрудника; при нескольких сотрудниках KPI является совместным"
    columns = [
        "ID сотрудника",
        "Ключ эпизода",
        "Проект",
        "Месяц",
        "Итоговый KPI",
        "Рабочий порог KPI",
        "KPI 2",
        "Сервис факт",
        "OSA факт",
        "PICOS факт",
        "Покрытие факт",
        "Покрытие план",
        "PICOS СВ факт",
        "PICOS СВ план",
        "Количество подтверждённых визитов RTM",
        "Количество ТТ с клиентским KPI",
        "Окончательный расчёт месяца",
        "Дата формирования показателя",
        "Регион",
        "Руководитель в месяце",
        "Метод распределения",
    ]
    return grouped[columns].sort_values(["Месяц", "ID сотрудника"])


def fraud_reason(row: pd.Series, reason_columns: list[str]) -> str:
    reasons = []
    comment = row.get("Фальсификация: комментарий")
    if pd.notna(comment) and str(comment).strip():
        reasons.append(str(comment).strip())
    for column in reason_columns:
        value = row.get(column)
        if pd.notna(value) and bool(value):
            reasons.append(column.replace("Фальсификация причины: ", ""))
    return "; ".join(dict.fromkeys(reasons)) if reasons else pd.NA


def build_okk_checks(facts, episodes: pd.DataFrame, maps) -> pd.DataFrame:
    okk = facts["okk"].copy().reset_index(drop=True)
    reason_columns = [column for column in okk.columns if column.startswith("Фальсификация причины:")]
    standard_columns = [
        column
        for column in okk.columns
        if column.startswith("Правила фотографирования:")
        or column.startswith("Стандарты:")
        or column.startswith("check_")
    ]
    okk["ID сотрудника raw"] = okk["ID мерчендайзера"].astype("string").str.strip()
    okk["ID сотрудника"] = okk["ID сотрудника raw"].map(maps["employee"]).fillna("EMP_UNKNOWN")
    okk["Проект"] = PROJECT_NAME
    okk["ID торговой точки"] = okk["Код ТТ"].map(lambda value: anon(value, maps["tt"], "TT_UNKNOWN"))
    okk["ID визита"] = okk.apply(
        lambda row: stable_hash(
            row["ID сотрудника"], row["Дата визита"], row["ID торговой точки"], row.name, prefix="VISIT"
        ),
        axis=1,
    )
    duplicate_key = ["ID мерчендайзера", "Дата визита", "Код ТТ", "Качество визита", "Флаг фальсификации"]
    okk["Дубль проверки"] = okk.duplicated(duplicate_key, keep="first")
    okk["Фрод-сигнал"] = okk["Флаг фальсификации"]
    okk["Причина фрод-сигнала"] = okk.apply(lambda row: fraud_reason(row, reason_columns), axis=1)
    okk["Количество проверенных элементов"] = okk[standard_columns].notna().sum(axis=1)
    okk["Ключ эпизода"] = okk.apply(
        lambda row: assign_episode(row["ID сотрудника raw"], row["Дата визита"], episodes), axis=1
    )
    base_columns = [
        "ID сотрудника",
        "Ключ эпизода",
        "Проект",
        "Дата визита",
        "ID визита",
        "ID торговой точки",
        "Качество визита",
        "% качества PICoS",
        "Качество 1-ого фотоаудита",
        "% наличия товара на полке",
        "% наличия PICoS",
        "Фрод-сигнал",
        "Причина фрод-сигнала",
        "Количество проверенных элементов",
        "Дубль проверки",
    ]
    return okk[base_columns + standard_columns].rename(
        columns={
            "Дата визита": "Дата проверки",
            "Качество визита": "Итоговый ОКК",
            "% качества PICoS": "PICOS",
            "Качество 1-ого фотоаудита": "Фотоаудит",
            "% наличия товара на полке": "Доступность",
        }
    ).sort_values(["Дата проверки", "ID сотрудника"])


def _normalize_rtm_bool(series: pd.Series) -> pd.Series:
    return (
        series.astype("string")
        .str.strip()
        .str.lower()
        .map(
            {
                "true": True,
                "1": True,
                "да": True,
                "yes": True,
                "false": False,
                "0": False,
                "нет": False,
                "no": False,
            }
        )
        .astype("boolean")
    )


def _load_rtm_rows() -> tuple[pd.DataFrame, pd.DataFrame]:
    required = [
        "route_date",
        "route_name",
        "employee_id",
        "shop_code",
        "agg_visit_id",
        "visit_date",
        "visit_id",
        "visit_status",
        "visit_cancel_reason_status",
        "is_planned",
        "is_complete",
        "is_confirmed",
    ]
    frames = []
    audit_rows = []
    rtm_root = RAW / "kpi" / "RTM"
    for source in sorted(rtm_root.rglob("*.xlsx")):
        workbook = load_workbook(source, read_only=True, data_only=True)
        worksheet = workbook.active
        worksheet.reset_dimensions()
        rows = worksheet.iter_rows(values_only=True)
        filter_text = str(next(rows)[0] or "")
        next(rows, None)
        headers = list(next(rows))
        while headers and headers[-1] is None:
            headers.pop()
        missing = [column for column in required if column not in headers]
        if missing:
            raise ValueError(f"RTM {source.name}: отсутствуют поля {missing}")
        positions = {column: headers.index(column) for column in required}
        records = []
        for row in rows:
            records.append({column: row[position] for column, position in positions.items()})
        frame = pd.DataFrame(records)
        frame["Файл RTM"] = source.relative_to(rtm_root).as_posix()
        frame["Фильтр выгрузки RTM"] = filter_text
        frame["Дата визита"] = pd.to_datetime(frame["visit_date"], errors="coerce").dt.normalize()
        observed = frame["Дата визита"].dropna()
        audit_rows.append(
            {
                "Файл RTM": source.relative_to(rtm_root).as_posix(),
                "Фильтр выгрузки": filter_text,
                "Строк": len(frame),
                "Минимальная дата": observed.min() if not observed.empty else pd.NaT,
                "Максимальная дата": observed.max() if not observed.empty else pd.NaT,
                "Возможен лимит выгрузки": len(frame) >= 150_000,
            }
        )
        frames.append(frame)
    if not frames:
        raise FileNotFoundError("В data/raw/kpi/RTM нет файлов RTM")
    result = pd.concat(frames, ignore_index=True)
    for column in ["is_complete", "is_confirmed"]:
        result[column] = _normalize_rtm_bool(result[column])

    missing_date = result["Дата визита"].isna()
    missing_flags = result["is_complete"].isna() | result["is_confirmed"].isna()
    quality_rows = []
    for label, mask in [
        ("Удалено: нет фактической даты визита", missing_date),
        ("Удалено: пустой флаг выполнения/подтверждения", missing_flags & ~missing_date),
    ]:
        if mask.any():
            quality_rows.extend(
                result.loc[mask]
                .groupby("Файл RTM", dropna=False)
                .size()
                .reset_index(name="Строк")
                .assign(Проверка=label)
                .to_dict("records")
            )

    result = result[~missing_date & ~missing_flags].copy()
    return result, pd.DataFrame([*audit_rows, *quality_rows])


def _validate_rtm_periods(rows: pd.DataFrame):
    dates = rows["Дата визита"].dropna()
    problems = []
    for year_month in RTM_REQUIRED_MONTHS:
        year, month = divmod(year_month, 100)
        month_start = pd.Timestamp(year, month, 1)
        month_end = month_start + pd.offsets.MonthEnd(0)
        month_dates = dates[dates.dt.to_period("M").eq(month_start.to_period("M"))]
        if month_dates.empty:
            problems.append(f"{month_start:%m.%Y}: данных нет")
            continue
        if month_dates.min() > month_start + pd.Timedelta(days=2):
            problems.append(f"{month_start:%m.%Y}: начало только {month_dates.min():%d.%m.%Y}")
        if month_dates.max() < month_end - pd.Timedelta(days=2):
            problems.append(f"{month_start:%m.%Y}: окончание уже {month_dates.max():%d.%m.%Y}")
    if problems:
        raise ValueError("Неполный период RTM 05–07.2026: " + "; ".join(problems))


def _build_rtm_employee_bridge(rows: pd.DataFrame, episodes: pd.DataFrame) -> pd.DataFrame:
    dim = pd.read_parquet(OUT / "dim_employees.parquet")
    hr = pd.read_parquet(OUT / "fact_hr_registry.parquet")
    login_map, _ = load_login_employee_map(RAW / "kpi" / "Логины", dim, hr)
    bridge = login_map.rename(
        columns={"Код RTM": "employee_id", "ID сотрудника": "ID сотрудника raw"}
    )
    episode_ids = set(episodes["ID сотрудника raw"].dropna().astype(str))
    bridge = bridge[bridge["ID сотрудника raw"].astype(str).isin(episode_ids)].copy()
    return bridge[["YearMonth", "employee_id", "ID сотрудника raw"]].drop_duplicates()


def build_field_visits(
    episodes: pd.DataFrame,
    maps,
    rows: pd.DataFrame | None = None,
    audit: pd.DataFrame | None = None,
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    if rows is None or audit is None:
        rows, audit = _load_rtm_rows()
        rows = rows[rows["Дата визита"].between(RTM_START, RTM_END)].copy()
        _validate_rtm_periods(rows)
    rows["employee_id"] = rows["employee_id"].astype("string").str.strip()
    rows["employee_id"] = rows["employee_id"].str.replace(r"\.0+$", "", regex=True)
    rows["YearMonth"] = rows["Дата визита"].dt.year * 100 + rows["Дата визита"].dt.month
    rows["route_name"] = rows["route_name"].astype("string").str.strip()
    rows["visit_id"] = rows["visit_id"].astype("string").str.strip()
    rows["agg_visit_id"] = rows["agg_visit_id"].astype("string").str.strip()
    rows["Ключ визита RTM"] = rows["visit_id"].combine_first(rows["agg_visit_id"])
    rows = rows[rows["Ключ визита RTM"].notna()].copy()
    visits = (
        rows.groupby("Ключ визита RTM", dropna=False)
        .agg(
            YearMonth=("YearMonth", "min"),
            employee_id=("employee_id", first_notna),
            Дата_визита=("Дата визита", "min"),
            Маршрут=("route_name", mode_or_first),
            Код_ТТ=("shop_code", first_notna),
            Визит_запланирован=("is_planned", "max"),
            Визит_выполнен=("is_complete", "max"),
            Визит_подтверждён=("is_confirmed", "max"),
            Причина_отмены=("visit_cancel_reason_status", first_notna),
        )
        .reset_index()
    )
    visits = visits[visits["Визит_выполнен"].eq(True) & visits["Визит_подтверждён"].eq(True)].copy()
    bridge = _build_rtm_employee_bridge(rows, episodes)
    coverage = (
        bridge.groupby("ID сотрудника raw")["employee_id"]
        .nunique()
        .reset_index(name="Количество связанных RTM ID")
    )
    coverage["ID сотрудника"] = coverage["ID сотрудника raw"].map(maps["employee"]).fillna("EMP_UNKNOWN")
    coverage["RTM сопоставлен"] = True
    visits = visits.merge(bridge, on=["YearMonth", "employee_id"], how="left")
    visits = visits[visits["ID сотрудника raw"].notna()].copy()
    visits = visits.merge(
        episodes[
            [
                "ID сотрудника raw",
                "Дата приёма на проект",
                "Дата увольнения с проекта",
                "Ключ эпизода",
                "Супервайзер",
            ]
        ],
        on="ID сотрудника raw",
        how="inner",
    )
    visits = visits[
        visits["Дата_визита"].ge(visits["Дата приёма на проект"])
        & (
            visits["Дата увольнения с проекта"].isna()
            | visits["Дата_визита"].le(visits["Дата увольнения с проекта"])
        )
    ].copy()
    visits = visits.sort_values(["Ключ визита RTM", "Дата приёма на проект"]).drop_duplicates(
        "Ключ визита RTM", keep="last"
    )
    visits["ID сотрудника"] = visits["ID сотрудника raw"].map(maps["employee"]).fillna("EMP_UNKNOWN")
    visits["Проект"] = PROJECT_NAME
    visits["Дата визита"] = visits["Дата_визита"]
    visits["ID визита"] = visits["Ключ визита RTM"].map(lambda value: stable_hash(value, prefix="RTM_VISIT"))
    visits["ID торговой точки"] = visits["Код_ТТ"].map(lambda value: anon(value, maps["tt"], "TT_UNKNOWN"))
    visits["Визит запланирован"] = visits["Визит_запланирован"]
    visits["Визит выполнен"] = visits["Визит_выполнен"]
    visits["Визит подтверждён"] = visits["Визит_подтверждён"]
    visits["Самостоятельный / совместный"] = "Нет признака в RTM"
    visits["Супервайзер"] = visits["Супервайзер"].fillna("SV_UNKNOWN")
    visits["Причина отмены"] = visits["Причина_отмены"]
    visits["Правило полевого выхода"] = "Первый визит RTM после найма: is_complete = TRUE и is_confirmed = TRUE"
    columns = [
        "ID сотрудника",
        "Ключ эпизода",
        "Проект",
        "Дата визита",
        "ID визита",
        "ID торговой точки",
        "Визит запланирован",
        "Визит выполнен",
        "Визит подтверждён",
        "Самостоятельный / совместный",
        "Супервайзер",
        "Маршрут",
        "Причина отмены",
        "Правило полевого выхода",
    ]
    audit["Используется в расчёте"] = audit["Максимальная дата"].ge(RTM_START) & audit["Минимальная дата"].le(RTM_END)
    return (
        visits[columns].sort_values(["Дата визита", "ID сотрудника"]),
        audit,
        coverage[["ID сотрудника", "RTM сопоставлен", "Количество связанных RTM ID"]],
    )


def learning_pct_as_of(learning: pd.DataFrame, employee_id: str, checkpoint: pd.Timestamp):
    courses = learning[
        learning["ID сотрудника"].eq(employee_id) & learning["Обязательный"].eq(True)
    ].copy()
    assigned = pd.to_datetime(courses["Дата назначения"], errors="coerce").dt.normalize()
    completed_date = pd.to_datetime(courses["Дата завершения"], errors="coerce").dt.normalize()
    courses = courses[assigned.le(checkpoint) | (assigned.isna() & completed_date.le(checkpoint))]
    courses = courses.sort_values(["Номер курса", "Дата завершения"]).drop_duplicates("Номер курса", keep="last")
    if courses.empty:
        return pd.NA
    completed = courses[
        courses["Успешно завершён"].eq(True)
        & pd.to_datetime(courses["Дата завершения"], errors="coerce").dt.normalize().le(checkpoint)
    ]["Номер курса"].nunique()
    return completed / courses["Номер курса"].nunique()


def month_result(kpi: pd.DataFrame, okk: pd.DataFrame, employee_id: str, month_start: pd.Timestamp):
    kpi_row = kpi[
        kpi["ID сотрудника"].eq(employee_id) & kpi["Месяц"].eq(month_start)
    ]
    month_end = month_start + pd.offsets.MonthEnd(0)
    okk_rows = okk[
        okk["ID сотрудника"].eq(employee_id)
        & okk["Дата проверки"].between(month_start, month_end)
        & ~okk["Дубль проверки"].eq(True)
    ]
    return {
        "KPI": kpi_row["Итоговый KPI"].mean() if not kpi_row.empty else pd.NA,
        "ОКК": okk_rows["Итоговый ОКК"].mean() if not okk_rows.empty else pd.NA,
        "Фрод": okk_rows["Фрод-сигнал"].mean() if not okk_rows.empty else pd.NA,
        "Проверки": len(okk_rows),
    }


def build_may_cohort(episodes, learning, kpi, okk, field_visits, rtm_coverage):
    cohort = episodes[
        episodes["Дата приёма на проект"].between(MAY_START, MAY_END)
        & episodes["Должность"].astype("string").str.contains("мерч", case=False, na=False)
    ].copy()
    first_visits = (
        field_visits.merge(
            cohort[["Ключ эпизода", "Дата приёма на проект"]], on="Ключ эпизода", how="inner"
        )
    )
    first_visits = first_visits[first_visits["Дата визита"].ge(first_visits["Дата приёма на проект"])]
    first_visits = first_visits.groupby("Ключ эпизода")["Дата визита"].min().rename("Дата первого полевого выхода")
    cohort = cohort.merge(first_visits, on="Ключ эпизода", how="left")
    cohort = cohort.merge(rtm_coverage, on="ID сотрудника", how="left")
    cohort["RTM сопоставлен"] = cohort["RTM сопоставлен"].fillna(False).astype(bool)
    cohort["Дней до поля"] = (cohort["Дата первого полевого выхода"] - cohort["Дата приёма на проект"]).dt.days
    cohort["Вышел в поле за 7 дней"] = cohort["Дней до поля"].between(0, 7).astype("boolean")
    cohort["Вышел в поле за 14 дней"] = cohort["Дней до поля"].between(0, 14).astype("boolean")
    cohort.loc[~cohort["RTM сопоставлен"], ["Вышел в поле за 7 дней", "Вышел в поле за 14 дней"]] = pd.NA
    cohort["Статус полевого наблюдения"] = np.select(
        [
            ~cohort["RTM сопоставлен"],
            cohort["Вышел в поле за 14 дней"].fillna(False),
            cohort["Дата первого полевого выхода"].notna(),
        ],
        ["Недостаточно данных RTM", "Вышел за 14 дней", "Вышел позднее 14 дней"],
        default="Нет подтверждённого визита RTM",
    )

    detail_rows = []
    for _, row in cohort.iterrows():
        employee_id = row["ID сотрудника"]
        hire = row["Дата приёма на проект"]
        training_pre = learning_pct_as_of(learning, employee_id, hire - pd.Timedelta(days=1))
        training_14 = learning_pct_as_of(learning, employee_id, hire + pd.Timedelta(days=14))
        training_june = learning_pct_as_of(learning, employee_id, JUNE_END)
        training_july = learning_pct_as_of(learning, employee_id, JULY_END)
        june = month_result(kpi, okk, employee_id, pd.Timestamp("2026-06-01"))
        july = month_result(kpi, okk, employee_id, pd.Timestamp("2026-07-01"))
        early = okk[
            okk["ID сотрудника"].eq(employee_id)
            & okk["Дата проверки"].between(hire, hire + pd.Timedelta(days=14))
            & ~okk["Дубль проверки"].eq(True)
        ]
        early_okk = early["Итоговый ОКК"].mean() if not early.empty else pd.NA
        early_fraud = bool(early["Фрод-сигнал"].any()) if not early.empty else pd.NA
        june_ready = bool(
            pd.notna(training_june)
            and training_june >= TRAINING_TARGET
            and pd.notna(row["Дата первого полевого выхода"])
            and row["Дата первого полевого выхода"] <= JUNE_END
            and pd.notna(june["KPI"])
            and june["KPI"] >= KPI_TARGET
            and pd.notna(june["ОКК"])
            and june["ОКК"] >= OKK_TARGET
            and june["Фрод"] <= FRAUD_MAX
        )
        july_ready = bool(
            pd.notna(training_july)
            and training_july >= TRAINING_TARGET
            and pd.notna(row["Дата первого полевого выхода"])
            and row["Дата первого полевого выхода"] <= JULY_END
            and pd.notna(july["KPI"])
            and july["KPI"] >= KPI_TARGET
            and pd.notna(july["ОКК"])
            and july["ОКК"] >= OKK_TARGET
            and july["Фрод"] <= FRAUD_MAX
        )
        if pd.isna(training_july) or training_july < TRAINING_TARGET:
            delay = "Обучение"
        elif pd.isna(row["Дата первого полевого выхода"]):
            delay = "Выход в поле"
        elif pd.isna(july["KPI"]):
            delay = "Нет KPI"
        elif july["KPI"] < KPI_TARGET:
            delay = "KPI"
        elif pd.isna(july["ОКК"]):
            delay = "Нет ОКК"
        elif july["ОКК"] < OKK_TARGET or july["Фрод"] > FRAUD_MAX:
            delay = "Качество"
        else:
            delay = "Готов"

        signals = []
        signal_dates = []
        if pd.notna(row["Вышел в поле за 14 дней"]) and not bool(row["Вышел в поле за 14 дней"]):
            signals.append("Нет выхода в поле за 14 дней")
            signal_dates.append(hire + pd.Timedelta(days=14))
        if pd.isna(training_14) or training_14 < 0.90:
            signals.append("Обучение <90% на 14-й день")
            signal_dates.append(hire + pd.Timedelta(days=14))
        if pd.isna(early_okk):
            signals.append("Нет ОКК в первые 14 дней")
            signal_dates.append(hire + pd.Timedelta(days=14))
        elif early_okk < OKK_TARGET:
            signals.append("ОКК <50% в первые 14 дней")
            signal_dates.append(early["Дата проверки"].min())
        if early_fraud is True:
            signals.append("Фрод в первые 14 дней")
            signal_dates.append(early.loc[early["Фрод-сигнал"].eq(True), "Дата проверки"].min())

        detail_rows.append(
            {
                "Ключ эпизода": row["Ключ эпизода"],
                "ID сотрудника": employee_id,
                "Регион": row["Регион"],
                "Супервайзер": row["Супервайзер"],
                "Территориальный менеджер": row["Территориальный менеджер"],
                "Дата приёма": hire,
                "Дата увольнения": row["Дата увольнения с проекта"],
                "Тип выхода": row["Тип выхода"],
                "Кадровое состояние": row["Кадровое состояние"],
                "Активен в USERS": row["Активен в USERS"],
                "Увольнение подтверждено": row["Увольнение подтверждено"],
                "Конфликт кадровой даты": row["Конфликт кадровой даты"],
                "Подготовлен до выхода >=90%": pd.notna(training_pre) and training_pre >= TRAINING_TARGET,
                "Предобучение закрыто %": training_pre,
                "Обучение на 14-й день %": training_14,
                "Обучение на 30 июня %": training_june,
                "Обучение на 31 июля %": training_july,
                "Дата первого полевого выхода": row["Дата первого полевого выхода"],
                "RTM сопоставлен": row["RTM сопоставлен"],
                "Статус полевого наблюдения": row["Статус полевого наблюдения"],
                "Дней до поля": row["Дней до поля"],
                "Вышел в поле за 7 дней": row["Вышел в поле за 7 дней"],
                "Вышел в поле за 14 дней": row["Вышел в поле за 14 дней"],
                "KPI июня": june["KPI"],
                "ОКК июня": june["ОКК"],
                "Фрод июня": june["Фрод"],
                "Проверок июня": june["Проверки"],
                "KPI июля": july["KPI"],
                "ОКК июля": july["ОКК"],
                "Фрод июля": july["Фрод"],
                "Проверок июля": july["Проверки"],
                "Цель достигнута к концу июня": june_ready,
                "Цель достигнута только к июлю": (not june_ready) and july_ready,
                "Результат сохранён в июне и июле": june_ready and july_ready,
                "Готов на 31 июля": july_ready,
                "Дата готовности": JUNE_END if june_ready else JULY_END if july_ready else pd.NaT,
                "Дней до готовности": (
                    (JUNE_END - hire).days if june_ready else (JULY_END - hire).days if july_ready else pd.NA
                ),
                "Где задержался путь": delay,
                "Сигналы первых 14 дней": ", ".join(signals) if signals else "Ранних сигналов нет",
                "Количество ранних сигналов": len(signals),
                "Дата первого раннего сигнала": min(signal_dates) if signal_dates else pd.NaT,
                "Дней раннего обнаружения до 31 июля": (JULY_END - min(signal_dates)).days if signal_dates else pd.NA,
                "ОКК первых 14 дней": early_okk,
                "Фрод первых 14 дней": early_fraud,
            }
        )
    detail = pd.DataFrame(detail_rows)

    metrics = [
        ("Кадровых эпизодов в мае", len(detail), "эпизодов", "Все эпизоды МЕ с датой выхода в мае"),
        (
            "Подтверждённых увольнений к 31 июля",
            (detail["Увольнение подтверждено"].eq(True) & detail["Дата увольнения"].le(JULY_END)).sum(),
            "чел.",
            "Есть дата, кадровое состояние 'Увольнение' и сотрудник не активен в USERS либо есть повторный найм",
        ),
        (
            "Сверхранний отток до 14 дней",
            (
                detail["Увольнение подтверждено"].eq(True)
                & detail["Дата увольнения"].le(JULY_END)
                & (detail["Дата увольнения"] - detail["Дата приёма"]).dt.days.le(14)
            ).sum(),
            "чел.",
            "Подтверждённое увольнение не позднее 14-го дня",
        ),
        ("Сопоставлены с RTM", detail["RTM сопоставлен"].sum(), "чел.", "Есть однозначная связка RTM с кадровым эпизодом"),
        ("Фактически вышли в поле", detail["Дата первого полевого выхода"].notna().sum(), "чел.", "Есть выполненный и подтверждённый визит RTM после выхода"),
        ("Подготовлены до выхода", detail["Подготовлен до выхода >=90%"].sum(), "чел.", "До даты выхода успешно закрыто >=90% обязательных курсов"),
        ("Вышли в поле за 7 дней", detail["Вышел в поле за 7 дней"].sum(), "чел.", "Первый подтверждённый визит не позднее 7-го дня"),
        ("Вышли в поле за 14 дней", detail["Вышел в поле за 14 дней"].sum(), "чел.", "Первый подтверждённый визит не позднее 14-го дня"),
        ("Достигли цели к концу июня", detail["Цель достигнута к концу июня"].sum(), "чел.", "Обучение, поле, KPI, ОКК и фрод в норме"),
        ("Достигли цели только к июлю", detail["Цель достигнута только к июлю"].sum(), "чел.", "Не были готовы в июне, но готовы в июле"),
        ("Сохранили результат в июне и июле", detail["Результат сохранён в июне и июле"].sum(), "чел.", "Целевой результат в обоих месяцах"),
    ]
    risk_cases = detail[(detail["Готов на 31 июля"].eq(False)) & detail["Дата первого раннего сигнала"].notna()]
    metrics.append(
        (
            "Медиана раннего обнаружения риска",
            risk_cases["Дней раннего обнаружения до 31 июля"].median(),
            "дней раньше",
            "Разница между первым сигналом первых 14 дней и итогом июля",
        )
    )
    summary = pd.DataFrame(metrics, columns=["Метрика", "Значение", "Единица", "Правило"])
    count_mask = summary["Единица"].isin(["чел.", "эпизодов"])
    summary["Доля майской когорты"] = pd.NA
    summary.loc[count_mask, "Доля майской когорты"] = summary.loc[count_mask, "Значение"] / len(detail)
    summary["Доля наблюдаемой RTM выборки"] = pd.NA
    rtm_count = int(detail["RTM сопоставлен"].sum())
    rtm_metrics = summary["Метрика"].isin(
        ["Фактически вышли в поле", "Вышли в поле за 7 дней", "Вышли в поле за 14 дней"]
    )
    if rtm_count:
        summary.loc[rtm_metrics, "Доля наблюдаемой RTM выборки"] = (
            summary.loc[rtm_metrics, "Значение"] / rtm_count
        )
    ready_days = pd.to_numeric(detail["Дней до готовности"], errors="coerce").dropna()
    summary = pd.concat(
        [
            summary,
            pd.DataFrame(
                [
                    {
                        "Метрика": "Медианное время от выхода до готовности",
                        "Значение": ready_days.median() if not ready_days.empty else pd.NA,
                        "Единица": "дней",
                        "Правило": "До первой контрольной даты, когда одновременно выполнены обучение, поле, KPI и качество",
                         "Доля майской когорты": pd.NA,
                         "Доля наблюдаемой RTM выборки": pd.NA,
                    }
                ]
            ),
        ],
        ignore_index=True,
    )
    bottlenecks = detail.groupby("Где задержался путь")["Ключ эпизода"].nunique().reset_index(name="Сотрудников")
    bottlenecks["Доля когорты"] = bottlenecks["Сотрудников"] / len(detail)

    stage_masks = [
        ("Вышли на проект в мае", pd.Series(True, index=detail.index)),
        ("Закрыли обязательное обучение к 31 июля", detail["Обучение на 31 июля %"].ge(TRAINING_TARGET)),
        ("После обучения вышли в поле", detail["Обучение на 31 июля %"].ge(TRAINING_TARGET) & detail["Дата первого полевого выхода"].notna()),
        (
            "После выхода достигли KPI июля",
            detail["Обучение на 31 июля %"].ge(TRAINING_TARGET)
            & detail["Дата первого полевого выхода"].notna()
            & detail["KPI июля"].ge(KPI_TARGET),
        ),
        (
            "Достигли полной готовности в июле",
            detail["Готов на 31 июля"].eq(True),
        ),
        (
            "Сохранили результат в июне и июле",
            detail["Результат сохранён в июне и июле"].eq(True),
        ),
    ]
    funnel_rows = []
    previous_count = None
    for order, (stage, mask) in enumerate(stage_masks, start=1):
        count = int(mask.fillna(False).sum())
        funnel_rows.append(
            {
                "Порядок": order,
                "Этап": stage,
                "Сотрудников": count,
                "Доля от майской когорты": count / len(detail),
                "Конверсия от предыдущего этапа": count / previous_count if previous_count else pd.NA,
            }
        )
        previous_count = count
    funnel = pd.DataFrame(funnel_rows)
    return detail, summary, bottlenecks, funnel


def _month_text(month_start: pd.Timestamp) -> str:
    names = {
        1: "январь",
        2: "февраль",
        3: "март",
        4: "апрель",
        5: "май",
        6: "июнь",
        7: "июль",
        8: "август",
        9: "сентябрь",
        10: "октябрь",
        11: "ноябрь",
        12: "декабрь",
    }
    return f"{names[month_start.month]} {month_start.year}"


def _shared_tt_by_employee_month(field_visits: pd.DataFrame) -> pd.DataFrame:
    visits = field_visits.copy()
    visits["YearMonth"] = visits["Дата визита"].dt.year * 100 + visits["Дата визита"].dt.month
    people = (
        visits.groupby(["YearMonth", "ID торговой точки"])["ID сотрудника"]
        .nunique()
        .reset_index(name="Сотрудников на ТТ")
    )
    visits = visits.merge(people, on=["YearMonth", "ID торговой точки"], how="left")
    result = (
        visits.groupby(["YearMonth", "ID сотрудника"], dropna=False)
        .agg(
            **{
                "ТТ посещено": ("ID торговой точки", "nunique"),
                "Совместных ТТ": (
                    "ID торговой точки",
                    lambda values: values[visits.loc[values.index, "Сотрудников на ТТ"].gt(1)].nunique(),
                ),
            }
        )
        .reset_index()
    )
    result["Тип KPI"] = np.where(result["Совместных ТТ"].eq(0), "Персональный", "Совместный")
    return result


def build_expanded_cohorts(episodes, learning, kpi, okk, field_visits, rtm_coverage):
    cohort = episodes[
        episodes["Дата приёма на проект"].between(COHORT_START, COHORT_END)
        & episodes["Должность"].astype("string").str.contains("мерч", case=False, na=False)
    ].copy()
    cohort["Когорта"] = cohort["Дата приёма на проект"].dt.to_period("M").dt.to_timestamp()

    first_visits = field_visits.merge(
        cohort[["Ключ эпизода", "Дата приёма на проект"]], on="Ключ эпизода", how="inner"
    )
    first_visits = first_visits[first_visits["Дата визита"].ge(first_visits["Дата приёма на проект"])]
    first_visits = (
        first_visits.groupby("Ключ эпизода")["Дата визита"]
        .min()
        .rename("Дата первого полевого выхода")
    )
    cohort = cohort.merge(first_visits, on="Ключ эпизода", how="left")
    cohort = cohort.merge(rtm_coverage, on="ID сотрудника", how="left")
    cohort["RTM сопоставлен"] = cohort["RTM сопоставлен"].fillna(False).astype(bool)
    shared = _shared_tt_by_employee_month(field_visits)

    rows = []
    for _, employee in cohort.iterrows():
        employee_id = employee["ID сотрудника"]
        hire = employee["Дата приёма на проект"]
        cohort_month = employee["Когорта"]
        month_1 = cohort_month + pd.offsets.MonthBegin(1)
        month_2 = cohort_month + pd.offsets.MonthBegin(2)
        month_1_end = month_1 + pd.offsets.MonthEnd(0)
        month_2_end = month_2 + pd.offsets.MonthEnd(0)
        has_month_1 = month_1_end <= ANALYSIS_END
        has_month_2 = month_2_end <= ANALYSIS_END

        training_pre = learning_pct_as_of(learning, employee_id, hire - pd.Timedelta(days=1))
        training_14 = learning_pct_as_of(learning, employee_id, hire + pd.Timedelta(days=14))
        training_1 = learning_pct_as_of(learning, employee_id, month_1_end) if has_month_1 else pd.NA
        training_2 = learning_pct_as_of(learning, employee_id, month_2_end) if has_month_2 else pd.NA
        result_1 = month_result(kpi, okk, employee_id, month_1) if has_month_1 else {"KPI": pd.NA, "ОКК": pd.NA, "Фрод": pd.NA, "Проверки": 0}
        result_2 = month_result(kpi, okk, employee_id, month_2) if has_month_2 else {"KPI": pd.NA, "ОКК": pd.NA, "Фрод": pd.NA, "Проверки": 0}

        early = okk[
            okk["ID сотрудника"].eq(employee_id)
            & okk["Дата проверки"].between(hire, hire + pd.Timedelta(days=14))
            & ~okk["Дубль проверки"].eq(True)
        ].sort_values("Дата проверки")
        first_okk = early["Дата проверки"].min() if not early.empty else pd.NaT
        days_to_okk = (first_okk - hire).days if pd.notna(first_okk) else pd.NA
        early_okk = early["Итоговый ОКК"].mean() if not early.empty else pd.NA
        early_fraud = bool(early["Фрод-сигнал"].any()) if not early.empty else pd.NA

        employee_visits = field_visits[
            field_visits["Ключ эпизода"].eq(employee["Ключ эпизода"])
            & field_visits["Дата визита"].ge(hire)
        ]
        first_field = employee["Дата первого полевого выхода"]
        days_to_field = (first_field - hire).days if pd.notna(first_field) else pd.NA

        def ready(checkpoint, training, result, available):
            if not available:
                return pd.NA
            return bool(
                pd.notna(training)
                and training >= TRAINING_TARGET
                and pd.notna(first_field)
                and first_field <= checkpoint
                and pd.notna(result["KPI"])
                and result["KPI"] >= KPI_TARGET
                and pd.notna(result["ОКК"])
                and result["ОКК"] >= OKK_TARGET
                and pd.notna(result["Фрод"])
                and result["Фрод"] <= FRAUD_MAX
            )

        ready_1_base = ready(month_1_end, training_1, result_1, has_month_1)
        ready_2_base = ready(month_2_end, training_2, result_2, has_month_2)

        signals = []
        signal_dates = []
        if pd.isna(days_to_field) or days_to_field > 14:
            signals.append("Нет подтверждённого выхода в поле за 14 дней")
            signal_dates.append(hire + pd.Timedelta(days=14))
        if pd.isna(training_14) or training_14 < TRAINING_TARGET:
            signals.append("Обязательное обучение ниже 90% на 14-й день")
            signal_dates.append(hire + pd.Timedelta(days=14))
        if pd.isna(first_okk):
            signals.append("Нет ОКК в первые 14 дней")
            signal_dates.append(hire + pd.Timedelta(days=14))
        elif days_to_okk > 7:
            signals.append("Первая ОКК позже 7-го дня")
            signal_dates.append(first_okk)
        if pd.notna(early_okk) and early_okk < OKK_TARGET:
            signals.append("ОКК первых 14 дней ниже 50%")
            signal_dates.append(first_okk)
        if early_fraud is True:
            signals.append("Фрод-сигнал в первые 14 дней")
            signal_dates.append(early.loc[early["Фрод-сигнал"].eq(True), "Дата проверки"].min())

        shared_1 = shared[
            shared["YearMonth"].eq(month_1.year * 100 + month_1.month)
            & shared["ID сотрудника"].eq(employee_id)
        ]
        shared_2 = shared[
            shared["YearMonth"].eq(month_2.year * 100 + month_2.month)
            & shared["ID сотрудника"].eq(employee_id)
        ]
        type_1 = shared_1["Тип KPI"].iloc[0] if not shared_1.empty else "Нет RTM-визитов"
        type_2 = shared_2["Тип KPI"].iloc[0] if not shared_2.empty else "Нет RTM-визитов"
        ready_1 = (
            pd.NA if pd.isna(ready_1_base) else bool(ready_1_base and type_1 == "Персональный")
        )
        ready_2 = (
            pd.NA if pd.isna(ready_2_base) else bool(ready_2_base and type_2 == "Персональный")
        )
        if ready_1 is True:
            ready_date = month_1_end
        elif ready_2 is True:
            ready_date = month_2_end
        else:
            ready_date = pd.NaT

        if has_month_2 and ready_2 is True:
            status = "Готов и стабилен" if ready_1 is True else "Вышел на готовность ко 2-му месяцу"
        elif has_month_2:
            status = "Требуется поддержка"
        elif ready_1 is True:
            status = "Готов; устойчивость ещё не измерена"
        else:
            status = "Недостаточно полного окна"

        rows.append(
            {
                "Ключ эпизода": employee["Ключ эпизода"],
                "ID сотрудника": employee_id,
                "Когорта": cohort_month,
                "Регион": employee["Регион"],
                "Супервайзер": employee["Супервайзер"],
                "Территориальный менеджер": employee["Территориальный менеджер"],
                "Дата приёма": hire,
                "Дата увольнения": employee["Дата увольнения с проекта"],
                "Увольнение подтверждено": employee["Увольнение подтверждено"],
                "RTM сопоставлен": employee["RTM сопоставлен"],
                "Дата первого полевого выхода": first_field,
                "Дней до поля": days_to_field,
                "Вышел в поле за 7 дней": pd.notna(days_to_field) and 0 <= days_to_field <= 7,
                "Вышел в поле за 14 дней": pd.notna(days_to_field) and 0 <= days_to_field <= 14,
                "Дата первой ОКК": first_okk,
                "Дней до первой ОКК": days_to_okk,
                "ОКК в первые 7 дней": pd.notna(days_to_okk) and 0 <= days_to_okk <= 7,
                "ОКК в первые 14 дней": pd.notna(days_to_okk) and 0 <= days_to_okk <= 14,
                "ОКК первых 14 дней": early_okk,
                "Проверок ОКК первых 14 дней": len(early),
                "Предобучение %": training_pre,
                "Обучение на 14-й день %": training_14,
                "Месяц результата 1": month_1,
                "Обучение месяц 1 %": training_1,
                "KPI месяц 1": result_1["KPI"],
                "ОКК месяц 1": result_1["ОКК"],
                "Фрод месяц 1": result_1["Фрод"],
                "Проверок ОКК месяц 1": result_1["Проверки"],
                "Тип KPI месяц 1": type_1,
                "Готов месяц 1": ready_1,
                "Месяц результата 2": month_2 if has_month_2 else pd.NaT,
                "Обучение месяц 2 %": training_2,
                "KPI месяц 2": result_2["KPI"],
                "ОКК месяц 2": result_2["ОКК"],
                "Фрод месяц 2": result_2["Фрод"],
                "Проверок ОКК месяц 2": result_2["Проверки"],
                "Тип KPI месяц 2": type_2 if has_month_2 else pd.NA,
                "Готов месяц 2": ready_2,
                "Результат сохранён два месяца": ready_1 is True and ready_2 is True,
                "Дата готовности": ready_date,
                "Дней до готовности": (ready_date - hire).days if pd.notna(ready_date) else pd.NA,
                "Статус готовности": status,
                "Ранние сигналы": ", ".join(signals) if signals else "Ранних сигналов нет",
                "Дата первого сигнала": min(signal_dates) if signal_dates else pd.NaT,
                "План ОКК первых 14 дней": "Нет поля плана в источниках",
            }
        )

    detail = pd.DataFrame(rows)
    summary_rows = []
    for cohort_month, part in detail.groupby("Когорта", sort=True):
        full_second = part[part["Месяц результата 2"].notna()]
        ready_days = pd.to_numeric(part["Дней до готовности"], errors="coerce").dropna()
        summary_rows.append(
            {
                "Когорта": cohort_month,
                "Когорта текст": _month_text(cohort_month),
                "Новичков": len(part),
                "RTM сопоставлен %": part["RTM сопоставлен"].mean(),
                "Вышли в поле за 7 дней %": part["Вышел в поле за 7 дней"].mean(),
                "Вышли в поле за 14 дней %": part["Вышел в поле за 14 дней"].mean(),
                "Получили ОКК за 7 дней %": part["ОКК в первые 7 дней"].mean(),
                "Получили ОКК за 14 дней %": part["ОКК в первые 14 дней"].mean(),
                "Медиана дней до первой ОКК": pd.to_numeric(part["Дней до первой ОКК"], errors="coerce").median(),
                "Есть KPI месяц 1": part["KPI месяц 1"].notna().sum(),
                "Персональный KPI месяц 1": (part["Тип KPI месяц 1"] == "Персональный").sum(),
                "Совместный KPI месяц 1": (part["Тип KPI месяц 1"] == "Совместный").sum(),
                "Готовы месяц 1 %": part["Готов месяц 1"].eq(True).mean(),
                "Полное окно двух месяцев": len(full_second),
                "Готовы месяц 2 %": full_second["Готов месяц 2"].eq(True).mean() if len(full_second) else pd.NA,
                "Сохранили результат два месяца %": full_second["Результат сохранён два месяца"].mean() if len(full_second) else pd.NA,
                "Медиана дней до готовности": ready_days.median() if not ready_days.empty else pd.NA,
            }
        )
    summary = pd.DataFrame(summary_rows)

    okk_coverage = (
        detail.groupby("Когорта", sort=True)
        .agg(
            **{
                "Новичков": ("ID сотрудника", "nunique"),
                "ОКК за 7 дней": ("ОКК в первые 7 дней", "sum"),
                "ОКК за 14 дней": ("ОКК в первые 14 дней", "sum"),
                "Без ОКК за 14 дней": ("ОКК в первые 14 дней", lambda values: (~values).sum()),
                "Первая ОКК на 8–14 день": (
                    "Дней до первой ОКК",
                    lambda values: pd.to_numeric(values, errors="coerce").between(8, 14).sum(),
                ),
                "Медиана дней до первой ОКК": ("Дней до первой ОКК", "median"),
            }
        )
        .reset_index()
    )
    okk_coverage["Покрытие ОКК 7 дней %"] = okk_coverage["ОКК за 7 дней"] / okk_coverage["Новичков"]
    okk_coverage["Покрытие ОКК 14 дней %"] = okk_coverage["ОКК за 14 дней"] / okk_coverage["Новичков"]
    okk_coverage["Задержка ОКК 8–14 дней %"] = okk_coverage["Первая ОКК на 8–14 день"] / okk_coverage["Новичков"]
    okk_coverage["План проверок"] = "Не найден в источниках"

    eligible_1 = detail[detail["Месяц результата 1"].le(ANALYSIS_END)]
    eligible_2 = detail[detail["Месяц результата 2"].notna()]
    stages = [
        ("Вышли на проект", len(detail), len(detail)),
        ("Имеют надёжную связку RTM", int(detail["RTM сопоставлен"].sum()), len(detail)),
        ("Вышли в поле за 14 дней", int(detail["Вышел в поле за 14 дней"].sum()), len(detail)),
        ("Получили ОКК за 14 дней", int(detail["ОКК в первые 14 дней"].sum()), len(detail)),
        ("Есть KPI первого месяца", int(eligible_1["KPI месяц 1"].notna().sum()), len(eligible_1)),
        ("Готовы по итогам первого месяца", int(eligible_1["Готов месяц 1"].eq(True).sum()), len(eligible_1)),
        ("Готовы по итогам второго месяца", int(eligible_2["Готов месяц 2"].eq(True).sum()), len(eligible_2)),
        ("Сохранили результат два месяца", int(eligible_2["Результат сохранён два месяца"].sum()), len(eligible_2)),
    ]
    funnel = pd.DataFrame(
        [
            {"Порядок": index, "Этап": stage, "Сотрудников": count, "Наблюдаемая база": base, "Доля": count / base if base else pd.NA}
            for index, (stage, count, base) in enumerate(stages, start=1)
        ]
    )
    return detail, summary, okk_coverage, funnel


def signal_analysis(detail: pd.DataFrame) -> pd.DataFrame:
    observed = detail[
        detail[["KPI июля", "ОКК июля", "Фрод июля", "Обучение на 31 июля %"]].notna().all(axis=1)
    ].copy()
    signals = {
        "Нет выхода в поле за 14 дней": (
            ~observed["Вышел в поле за 14 дней"].eq(True),
            observed["RTM сопоставлен"].eq(True),
        ),
        "Обучение <90% на 14-й день": (
            observed["Обучение на 14-й день %"].isna() | observed["Обучение на 14-й день %"].lt(0.90),
            pd.Series(True, index=observed.index),
        ),
        "Нет ОКК в первые 14 дней": (
            observed["ОКК первых 14 дней"].isna(),
            pd.Series(True, index=observed.index),
        ),
        "ОКК <50% в первые 14 дней": (
            observed["ОКК первых 14 дней"].lt(OKK_TARGET),
            pd.Series(True, index=observed.index),
        ),
        "Фрод в первые 14 дней": (
            observed["Фрод первых 14 дней"].eq(True),
            pd.Series(True, index=observed.index),
        ),
    }
    rows = []
    for name, (mask, eligible) in signals.items():
        current = observed[eligible.fillna(False)].copy()
        mask = mask.reindex(current.index)
        outcome = current["Готов на 31 июля"].astype(int)
        mask = mask.fillna(False).astype(bool)
        yes = outcome[mask]
        no = outcome[~mask]
        correlation = mask.astype(int).corr(outcome) if mask.nunique() > 1 and outcome.nunique() > 1 else pd.NA
        rows.append(
            {
                "Сигнал первых 14 дней": name,
                "Наблюдаемая выборка": len(current),
                "С сигналом": int(mask.sum()),
                "Без сигнала": int((~mask).sum()),
                "Готовы к июлю при сигнале": yes.mean() if len(yes) else pd.NA,
                "Готовы к июлю без сигнала": no.mean() if len(no) else pd.NA,
                "Разница готовности": (yes.mean() - no.mean()) if len(yes) and len(no) else pd.NA,
                "Корреляция с готовностью июля": correlation,
                "Интерпретация": "Статистическая связь, не причинное влияние",
            }
        )
    return pd.DataFrame(rows).sort_values("Разница готовности")


def field_exit_analysis(detail: pd.DataFrame) -> pd.DataFrame:
    work = detail.copy()
    work["Дней до увольнения"] = (work["Дата увольнения"] - work["Дата приёма"]).dt.days
    work["Уволен к 31 июля"] = work["Увольнение подтверждено"].eq(True) & work[
        "Дата увольнения"
    ].le(JULY_END)
    work["Сверхранний отток"] = work["Уволен к 31 июля"] & work["Дней до увольнения"].le(14)
    full_window = work[~work["Сверхранний отток"]].copy()
    observed = full_window[full_window["RTM сопоставлен"].eq(True)].copy()
    rows = [
        {"Показатель": "Майская когорта", "Сотрудников": len(work), "Уволено": work["Уволен к 31 июля"].sum(), "Доля увольнений": work["Уволен к 31 июля"].mean(), "Комментарий": "Все кадровые эпизоды мая"},
        {"Показатель": "Сверхранний отток до 14 дней", "Сотрудников": int(work["Сверхранний отток"].sum()), "Уволено": int(work["Сверхранний отток"].sum()), "Доля увольнений": 1.0 if work["Сверхранний отток"].any() else pd.NA, "Комментарий": "Отдельная группа, исключена из сравнения полных 14 дней"},
        {"Показатель": "Полное окно 14 дней", "Сотрудников": len(full_window), "Уволено": full_window["Уволен к 31 июля"].sum(), "Доля увольнений": full_window["Уволен к 31 июля"].mean(), "Комментарий": "После исключения сверхраннего оттока"},
        {"Показатель": "Есть надёжная связка RTM", "Сотрудников": len(observed), "Уволено": observed["Уволен к 31 июля"].sum(), "Доля увольнений": observed["Уволен к 31 июля"].mean(), "Комментарий": "Наблюдаемая выборка для сравнения полевого выхода"},
    ]
    for field_value, label in [(True, "Вышел в поле за 14 дней"), (False, "Не вышел в поле за 14 дней")]:
        group = observed[observed["Вышел в поле за 14 дней"].eq(field_value)]
        rows.append(
            {
                "Показатель": label,
                "Сотрудников": len(group),
                "Уволено": group["Уволен к 31 июля"].sum(),
                "Доля увольнений": group["Уволен к 31 июля"].mean() if len(group) else pd.NA,
                "Комментарий": "Сравнивать доли нельзя при малой группе",
            }
        )
    return pd.DataFrame(rows)


def action_for_signal(reason: str):
    if "Фрод" in reason:
        return "Совместный контрольный визит и разбор фотоотчёта"
    if "ОКК" in reason:
        return "Совместный визит с СВ и разбор чек-листа ОКК"
    if "Обучение" in reason:
        return "Индивидуальный план закрытия обязательного обучения"
    if "поле" in reason:
        return "Назначить сопровождаемый первый визит"
    return "Разбор причин отклонения с сотрудником"


def build_actions(detail, okk, maps):
    candidates = detail[detail["Дата первого раннего сигнала"].notna()].sort_values(
        ["Количество ранних сигналов", "Дата первого раннего сигнала"], ascending=[False, True]
    )
    rows = []
    for _, employee in candidates.iterrows():
        signal_date = employee["Дата первого раннего сигнала"]
        no_action = len(rows) % 4 == 3
        action_date = pd.NaT if no_action else signal_date + pd.Timedelta(days=2)
        repeat_from = signal_date if no_action else action_date
        checks = okk[
            okk["ID сотрудника"].eq(employee["ID сотрудника"])
            & okk["Дата проверки"].gt(repeat_from)
            & okk["Дата проверки"].le(signal_date + pd.Timedelta(days=60))
            & ~okk["Дубль проверки"].eq(True)
        ].sort_values("Дата проверки")
        if checks.empty:
            continue
        repeat = checks.iloc[0]
        reason = employee["Сигналы первых 14 дней"].split(", ")[0]
        action = "Действие не проводилось" if no_action else action_for_signal(reason)
        success = bool(repeat["Итоговый ОКК"] >= OKK_TARGET and not repeat["Фрод-сигнал"])
        rows.append(
            {
                "ID сотрудника": employee["ID сотрудника"],
                "Ключ эпизода": employee["Ключ эпизода"],
                "Дата обнаружения проблемы": signal_date,
                "Кто обнаружил": "Система правил",
                "Причина сигнала": reason,
                "Фактическое действие": action,
                "Дата действия": action_date,
                "Ответственный": employee["Супервайзер"],
                "Дата повторной проверки": repeat["Дата проверки"],
                "Результат": "Вышел на норму" if success else "Не вышел на норму",
                "ОКК повторной проверки": repeat["Итоговый ОКК"],
                "Фрод повторной проверки": repeat["Фрод-сигнал"],
                "Статус действия": (
                    "Демонстрационный контрольный случай без действия"
                    if no_action
                    else "Демонстрационный сценарий; действие не подтверждено источником"
                ),
                "Статус результата": "Фактическая последующая проверка ОКК",
            }
        )
        if len(rows) == 20:
            break
    return pd.DataFrame(rows)


def confirmation_action(signals: str) -> str:
    actions = []
    if "Фрод" in signals:
        actions.append("совместный визит и разбор фото/фрод-сигнала")
    if "ОКК" in signals:
        actions.append("разбор ОКК по нарушениям")
    if "Обучение" in signals:
        actions.append("план закрытия обязательного обучения")
    if "Нет выхода в поле" in signals:
        actions.append("сопровождаемый первый визит")
    return "; ".join(actions) if actions else "контроль результата без дополнительного вмешательства"


def build_supervisor_confirmation(detail: pd.DataFrame) -> pd.DataFrame:
    priority_ids = [
        "EMP_A724803410",
        "EMP_6553A336CE",
        "EMP_5FB489B27F",
        "EMP_8B078AFE7A",
    ]
    selected_ids = [employee_id for employee_id in priority_ids if employee_id in set(detail["ID сотрудника"])]

    candidate_groups = [
        detail[
            detail["Цель достигнута только к июлю"].eq(True)
            & detail["Количество ранних сигналов"].gt(0)
        ].sort_values("Количество ранних сигналов", ascending=False),
        detail[
            detail["Результат сохранён в июне и июле"].eq(True)
        ].sort_values("Дней до готовности"),
        detail[
            detail["Готов на 31 июля"].eq(False)
            & detail["Количество ранних сигналов"].gt(0)
        ].sort_values("Количество ранних сигналов", ascending=False),
    ]
    for candidates in candidate_groups:
        for employee_id in candidates["ID сотрудника"]:
            if employee_id not in selected_ids:
                selected_ids.append(employee_id)
            if len(selected_ids) >= 8:
                break
        if len(selected_ids) >= 8:
            break

    rows = []
    for employee_id in selected_ids[:8]:
        employee = detail[detail["ID сотрудника"].eq(employee_id)].iloc[0]
        signals = employee["Сигналы первых 14 дней"]
        action = confirmation_action(signals)
        signal_date = employee["Дата первого раннего сигнала"]
        scenario_action_date = signal_date + pd.Timedelta(days=2) if pd.notna(signal_date) else pd.NaT
        if employee["Результат сохранён в июне и июле"]:
            trajectory = "Быстрый устойчивый выход"
        elif employee["Цель достигнута только к июлю"]:
            trajectory = "Восстановление к июлю"
        elif employee["Готов на 31 июля"]:
            trajectory = "Вышел на готовность"
        else:
            trajectory = "Требуется поддержка"
        rows.append(
            {
                "ID сотрудника": employee_id,
                "Тип траектории": trajectory,
                "Регион": employee["Регион"],
                "Супервайзер": employee["Супервайзер"],
                "Сигнал системы": signals,
                "Дата обнаружения системой": signal_date,
                "Сценарный ответ СВ: увидел проблему": "Да, увидел в карточке сотрудника",
                "Сценарная дата реакции СВ": signal_date,
                "Сценарное действие СВ": action,
                "Совместный визит": int("совместный визит" in action or "сопровождаемый" in action),
                "Разбор ОКК": int("ОКК" in action or "фрод" in action),
                "Обучение": int("обучения" in action),
                "Наставничество": int("сопровождаемый" in action or "обучения" in action),
                "Сценарная дата действия": scenario_action_date,
                "Подтверждено реальным СВ": "Нет — требуется интервью/подтверждение",
                "Фактическая дата действия": pd.NaT,
                "Фактический комментарий СВ": pd.NA,
                "KPI июня": employee["KPI июня"],
                "ОКК июня": employee["ОКК июня"],
                "KPI июля": employee["KPI июля"],
                "ОКК июля": employee["ОКК июля"],
                "Статус на 31 июля": "Готов" if employee["Готов на 31 июля"] else "Нужна поддержка",
                "Можно использовать как подтверждённый кейс": "Нет — до подтверждения действия СВ",
            }
        )
    return pd.DataFrame(rows)


def time_measurement() -> pd.DataFrame:
    rows = [
        {
            "Показатель": "Поиск информации по одному новичку — раньше",
            "Значение": 50,
            "Единица": "минут",
            "Статус": "Подтверждено внутренним замером",
            "Формула": pd.NA,
        },
        {
            "Показатель": "Просмотр готовой карточки — сейчас",
            "Значение": 6,
            "Единица": "минут",
            "Статус": "Подтверждено внутренним замером",
            "Формула": pd.NA,
        },
        {
            "Показатель": "Экономия на одной проверке",
            "Значение": 44,
            "Единица": "минут",
            "Статус": "Расчёт из подтверждённого замера",
            "Формула": "50 - 6",
        },
        {
            "Показатель": "Сокращение времени проверки",
            "Значение": 0.88,
            "Единица": "%",
            "Статус": "Расчёт из подтверждённого замера",
            "Формула": "(50 - 6) / 50",
        },
        {
            "Показатель": "Экономия на 10 проверках новичков",
            "Значение": 440,
            "Единица": "минут",
            "Статус": "Расчёт из подтверждённого замера",
            "Формула": "44 × 10",
        },
    ]
    return pd.DataFrame(rows)


def economics_assumptions(episodes: pd.DataFrame) -> pd.DataFrame:
    merch = episodes[episodes["Должность"].astype("string").str.contains("мерч", case=False, na=False)].copy()
    merch["Срок жизни, дней"] = (
        merch["Дата увольнения с проекта"] - merch["Дата приёма на проект"]
    ).dt.days
    lifetimes = merch.loc[merch["Срок жизни, дней"].ge(0), "Срок жизни, дней"]
    salary = 49_282
    annual_salary = salary * 12
    return pd.DataFrame(
        [
            {
                "Показатель": "Месячная зарплата мерчендайзера",
                "Минимум": salary,
                "Базовое значение": salary,
                "Максимум": salary,
                "Единица": "руб.",
                "Тип": "Рыночный ориентир",
                "Формула/основание": "Медиана предлагаемых зарплат в продуктах питания по данным hh.ru",
            },
            {
                "Показатель": "Стоимость подбора одного сотрудника",
                "Минимум": 15_000,
                "Базовое значение": 25_000,
                "Максимум": 35_000,
                "Единица": "руб.",
                "Тип": "Рыночный диапазон",
                "Формула/основание": "Диапазон массового подбора; заменить внутренним cost-per-hire",
            },
            {
                "Показатель": "Стоимость замены линейного сотрудника",
                "Минимум": annual_salary * 0.135,
                "Базовое значение": annual_salary * 0.1725,
                "Максимум": annual_salary * 0.21,
                "Единица": "руб.",
                "Тип": "Рыночная модель",
                "Формула/основание": "Годовая зарплата × 13,5–21%; не считать подтверждённой экономией",
            },
            {
                "Показатель": "Стоимость рабочего дня",
                "Минимум": salary / 21,
                "Базовое значение": salary / 21,
                "Максимум": salary / 21,
                "Единица": "руб./день",
                "Тип": "Расчёт",
                "Формула/основание": "Месячная зарплата / 21 рабочий день",
            },
            {
                "Показатель": "Медианный срок жизни уволившегося МЕ",
                "Минимум": lifetimes.median() if not lifetimes.empty else pd.NA,
                "Базовое значение": lifetimes.median() if not lifetimes.empty else pd.NA,
                "Максимум": lifetimes.median() if not lifetimes.empty else pd.NA,
                "Единица": "дней",
                "Тип": "Факт проекта",
                "Формула/основание": "Медиана: дата увольнения минус дата приёма",
            },
            {
                "Показатель": "Средний срок жизни уволившегося МЕ",
                "Минимум": lifetimes.mean() if not lifetimes.empty else pd.NA,
                "Базовое значение": lifetimes.mean() if not lifetimes.empty else pd.NA,
                "Максимум": lifetimes.mean() if not lifetimes.empty else pd.NA,
                "Единица": "дней",
                "Тип": "Факт проекта",
                "Формула/основание": "Среднее: дата увольнения минус дата приёма",
            },
        ]
    )


def economics_formulas() -> pd.DataFrame:
    return pd.DataFrame(
        [
            {
                "Эффект": "Экономия от предотвращённого раннего увольнения",
                "Формула": "Предотвращённые увольнения × стоимость замены одного сотрудника",
                "Что подставить": "Подтверждённое число предотвращённых увольнений и внутреннюю стоимость замены",
            },
            {
                "Эффект": "Экономия от сокращения времени до готовности",
                "Формула": "Сокращённые дни × число новичков × стоимость рабочего дня × доля потери производительности",
                "Что подставить": "Фактическое изменение Time-to-Ready и согласованную долю потери производительности",
            },
            {
                "Эффект": "Экономия времени руководителей",
                "Формула": "44 минуты × число проверок / 60 × стоимость часа руководителя",
                "Что подставить": "Количество проверок и внутреннюю стоимость часа СВ/ТМ",
            },
            {
                "Эффект": "Экономия годового потока найма",
                "Формула": "Экономия на одном новичке × число новичков за год",
                "Что подставить": "Подтверждённый эффект на сотрудника и годовое число выходов",
            },
        ]
    )


def economics_sources() -> pd.DataFrame:
    return pd.DataFrame(
        [
            {
                "Источник": "hh.ru — мониторинг операционного персонала",
                "Ссылка": "https://hh.ru/article/monitoring-operacionnogo-personala-spros-populyarnye-sfery-zanyatosti-i-zarplata",
                "Использовано": "49 282 руб. — медиана предлагаемой зарплаты мерчендайзера в продуктах питания",
            },
            {
                "Источник": "Atlas Projects — стоимость найма сотрудника",
                "Ссылка": "https://atlas-projects.ru/tpost/kak-poschitat-stoimost-najma-sotrudnika",
                "Использовано": "15–35 тыс. руб. — ориентир стоимости закрытия позиции массового персонала",
            },
            {
                "Источник": "Retail.ru — затраты на подбор и адаптацию в ритейле",
                "Ссылка": "https://www.retail.ru/articles/metodologiya-steadycontrol-kak-snizit-zatraty-na-podbor-i-adaptatsiyu-personala-v-riteyle/",
                "Использовано": "13,5% годового оклада — ориентир стоимости замены младшего линейного сотрудника",
            },
            {
                "Источник": "Современные проблемы экономики и менеджмента",
                "Ссылка": "https://sovman.ru/wp-content/uploads/2024/01/ss405.pdf",
                "Использовано": "21% годовой зарплаты — дополнительный верхний ориентир стоимости замены в России",
            },
        ]
    )


def client_kpi_targets(facts) -> pd.DataFrame:
    long = facts["kpi_long"].copy()
    long = long[long["YearMonth"].between(202601, 202607)].copy()
    summary = (
        long.groupby(["MonthStart", "YearMonth", "Блок KPI"], dropna=False)
        .agg(
            **{
                "Строк KPI": ("ТТ", "size"),
                "ТТ": ("ТТ", "nunique"),
                "Цель заполнена": ("Цель KPI", "count"),
                "Цель медиана": ("Цель KPI", "median"),
                "Цель минимум": ("Цель KPI", "min"),
                "Цель максимум": ("Цель KPI", "max"),
                "Факт медиана": ("Факт KPI", "median"),
                "Выполнение медиана": ("Выполнение KPI %", "median"),
            }
        )
        .reset_index()
    )
    summary["Тип цели"] = np.where(
        summary["Цель заполнена"].eq(summary["Строк KPI"]),
        "Клиентская цель блока заполнена",
        np.where(summary["Цель заполнена"].gt(0), "Клиентская цель заполнена частично", "Цели нет"),
    )
    summary["Комментарий"] = "Цель относится к KPI-блоку/ТТ и не является единой официальной планкой готовности новичка"
    return summary.sort_values(["MonthStart", "Блок KPI"])


def time_research_sources() -> pd.DataFrame:
    return pd.DataFrame(
        [
            {
                "Источник": "McKinsey Global Institute — The social economy",
                "Ссылка": "https://www.mckinsey.com/industries/technology-media-and-telecommunications/our-insights/capturing-business-value-with-social-technologies",
                "Что подтверждает": "Поиск и сбор информации занимает около 20% времени knowledge workers; единый поиск может сократить часть этих затрат.",
                "Ограничение": "Не измеряет именно работу супервайзера и проверку новичка.",
            },
            {
                "Источник": "Microsoft Customer Story — VSE",
                "Ссылка": "https://www.microsoft.com/en/customers/story/20328-vychodoslovenska-energetika-power-bi",
                "Что подтверждает": "До единой системы сотрудникам требовалось несколько минут и обращения к коллегам для поиска отдельных данных.",
                "Ограничение": "Другой процесс и отрасль.",
            },
            {
                "Источник": "Microsoft Customer Story — Centro de la Familia",
                "Ссылка": "https://www.microsoft.com/en/customers/story/23792-centro-de-la-familia-power-apps",
                "Что подтверждает": "Централизация данных дала кратное сокращение административного времени и перевела отчётность из часов в минуты.",
                "Ограничение": "Не является прямым benchmark для текущего проекта.",
            },
            {
                "Источник": "Power BI Community discussion — manual refresh workload",
                "Ссылка": "https://www.reddit.com/r/PowerBI/comments/1ovpr5m/is_it_normal_to_spend_12_hrs_everyday_doing_daily/",
                "Что подтверждает": "Пользователи описывают 1–2 часа ежедневной ручной работы с несколькими CSV/XLS/XLSX и отчётами.",
                "Ограничение": "Форумный пример, не статистическая выборка.",
            },
        ]
    )


def trajectory_row(label, row: pd.Series):
    return {
        "Тип траектории": label,
        "ID сотрудника": row["ID сотрудника"],
        "Регион": row["Регион"],
        "Дата выхода": row["Дата приёма"],
        "Предобучение %": row["Предобучение закрыто %"],
        "Дней до поля": row["Дней до поля"],
        "Ранние сигналы": row["Сигналы первых 14 дней"],
        "KPI июня": row["KPI июня"],
        "ОКК июня": row["ОКК июня"],
        "KPI июля": row["KPI июля"],
        "ОКК июля": row["ОКК июля"],
        "Статус июля": "Готов" if row["Готов на 31 июля"] else "Нужна помощь",
        "Комментарий": (
            "Быстро достиг и сохранил результат"
            if label == "Быстрый выход на эффективность"
            else "После раннего сигнала вышел на целевой уровень"
            if label == "Восстановление после раннего сигнала"
            else f"Основная задержка: {row['Где задержался путь']}"
        ),
    }


def build_trajectories(detail: pd.DataFrame) -> pd.DataFrame:
    rows = []
    fast = detail[detail["Цель достигнута к концу июня"].eq(True)].sort_values("Дней до поля")
    if not fast.empty:
        rows.append(trajectory_row("Быстрый выход на эффективность", fast.iloc[0]))
    recovery = detail[
        detail["Цель достигнута только к июлю"].eq(True)
        & detail["Количество ранних сигналов"].gt(0)
    ].sort_values("Количество ранних сигналов", ascending=False)
    if not recovery.empty:
        rows.append(trajectory_row("Восстановление после раннего сигнала", recovery.iloc[0]))
    help_needed = detail[detail["Готов на 31 июля"].eq(False)].sort_values(
        ["Количество ранних сигналов", "Проверок июля"], ascending=[False, False]
    )
    if not help_needed.empty:
        rows.append(trajectory_row("По-прежнему требуется помощь", help_needed.iloc[0]))
    return pd.DataFrame(rows)


def methodology() -> pd.DataFrame:
    rows = [
        ("Кадровый эпизод", "Одна строка = ID сотрудника + проект + дата приёма. Старые эпизоды сохраняются."),
        ("Тип выхода", "Первый эпизод = первичный найм; новый эпизод после увольнения = повторный найм; пересекающаяся запись = перевод/изменение."),
        ("Предобучение", "Курс назначен до даты приёма связанного кадрового эпизода."),
        ("Успешное обучение", "Курс завершён; для тестового курса балл не ниже порога каталога, для курса по прогрессу прогресс 100%."),
        ("Полевой визит", "Используется первый выполненный и подтверждённый визит RTM после даты найма."),
        ("Готовность", "Рабочее правило пилота: обязательное обучение >=90%, полевой выход подтверждён, KPI >=90%, ОКК >=50%, фрод <=20%. Это не клиентский SLA."),
        ("Готов к концу июня/июля", "Все условия готовности выполнены на соответствующую дату и есть показатели этого месяца."),
        ("Ранний сигнал", "В первые 14 дней: нет выхода в поле, обучение <90%, нет ОКК, первая ОКК позже 7-го дня, ОКК <50% или есть фрод."),
        ("Раннее обнаружение", "Дней между первым ранним сигналом и итоговой оценкой на 31 июля."),
        ("Подтверждённое увольнение", "Есть дата увольнения, кадровое состояние 'Увольнение' и сотрудник не активен в USERS либо после увольнения есть новый эпизод найма."),
        ("Наблюдаемость RTM", "Отсутствие визита считается сигналом только при однозначной связке RTM с сотрудником; иначе ставится недостаточно данных."),
        ("Дата готовности", "Первая доступная граница первого или второго полного месяца после выхода, когда одновременно выполнены все условия готовности."),
        ("Действия руководителя", "Действия сгенерированы как демонстрационный сценарий; повторные результаты взяты из фактического ОКК."),
    ]
    return pd.DataFrame(rows, columns=["Понятие", "Правило"])


def limitations() -> pd.DataFrame:
    rows = [
        ("Дата фактического выхода", "RTM не содержит отдельного признака самостоятельности", "Первый выполненный и подтверждённый визит RTM после найма"),
        ("Самостоятельный визит", "Признака нет", "Не интерпретируется; в файле указано 'нет признака'"),
        ("Полный журнал визитов", "RTM доступен за январь–июль 2026", "В расчёт входят только подтверждённые выполненные визиты"),
        ("Привязка RTM к сотруднику", "ID RTM отличается от ID USERS", "Код RTM связывается с месячным файлом логинов, затем точное ФИО — с ID USERS или кадровой истории; неоднозначные совпадения исключаются"),
        ("Полнота RTM июля", "В июльском файле ровно 150 002 строки — вероятен лимит выгрузки", "Июльские визиты считаются наблюдаемыми, но полноту выгрузки необходимо подтвердить"),
        ("Дата назначения обучения", "Отдельной даты нет", "Используется дата начала обучения"),
        ("Контур обучения", "Исходные учебные файлы содержат сотрудников за пределами проекта", "В отчёт включены сотрудники с кадровым эпизодом проекта"),
        ("Причина увольнения", "Нет в кадровом реестре", "Оставлена пустой"),
        ("Фактические действия СВ", "Журнал отсутствует", "Сгенерированы демонстрационные действия с явной маркировкой"),
        ("Совместный KPI ТТ", "В одной ТТ за месяц могли работать несколько сотрудников", "Такие результаты маркируются совместными; для строгого персонального кейса используются только ТТ с одним сотрудником"),
        ("Целевой KPI", "В клиентских файлах есть цели отдельных KPI-блоков, но нет утверждённой единой планки готовности новичка", "Используется мягкий рабочий порог пилота 90%; он не называется клиентским SLA"),
        ("ID визита", "Отдельный стабильный ID визита отсутствует", "Сформирован технический обезличенный ID проверки"),
        ("Историческая оргструктура", "HR содержит текущую/сопоставленную привязку", "Руководитель KPI берётся из KPI конкретного месяца"),
    ]
    return pd.DataFrame(rows, columns=["Поле", "Ограничение", "Принятое правило"])


def export_excel(tables: dict[str, pd.DataFrame]):
    REPORTS.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(OUTPUT, engine="openpyxl") as writer:
        for sheet, frame in tables.items():
            frame.to_excel(writer, sheet_name=sheet[:31], index=False)

    wb = load_workbook(OUTPUT)
    header_fill = PatternFill("solid", fgColor="0B4F3C")
    header_font = Font(color="FFFFFF", bold=True)
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        ws.row_dimensions[1].height = 34
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for index, header_cell in enumerate(ws[1], start=1):
            header = str(header_cell.value or "").lower()
            max_length = len(str(header_cell.value or ""))
            for cell in list(ws.columns)[index - 1][1:]:
                if cell.value is not None:
                    max_length = max(max_length, min(len(str(cell.value)), 55))
                if isinstance(cell.value, (int, float)) and any(marker in header for marker in ["%", "kpi", "окк", "фрод", "прогресс", "тест", "picos", "доступность", "фото"]):
                    cell.number_format = "0.0%"
                if hasattr(cell.value, "year") and any(marker in header for marker in ["дата", "месяц"]):
                    cell.number_format = "dd.mm.yyyy"
                cell.alignment = Alignment(vertical="top", wrap_text=True)
            ws.column_dimensions[get_column_letter(index)].width = max(11, min(max_length + 2, 42))
    wb.save(OUTPUT)


def main():
    rtm_rows, rtm_audit = _load_rtm_rows()
    rtm_rows = rtm_rows[rtm_rows["Дата визита"].between(RTM_START, RTM_END)].copy()
    _validate_rtm_periods(rtm_rows)
    facts = load_facts()
    raw_learning = load_raw_learning()
    maps = build_maps(facts, raw_learning)
    episodes = build_episodes(facts["hr"], maps)
    learning = build_learning(raw_learning, episodes, maps)
    kpi = build_kpi_monthly(facts, episodes, maps)
    okk = build_okk_checks(facts, episodes, maps)
    field_visits, rtm_audit, rtm_coverage = build_field_visits(
        episodes, maps, rtm_rows, rtm_audit
    )
    may_detail, may_summary, bottlenecks, funnel = build_may_cohort(
        episodes, learning, kpi, okk, field_visits, rtm_coverage
    )
    cohort_detail, cohort_summary, okk_coverage, cohort_funnel = build_expanded_cohorts(
        episodes, learning, kpi, okk, field_visits, rtm_coverage
    )
    signals = signal_analysis(may_detail)
    field_exit = field_exit_analysis(may_detail)
    actions = build_actions(may_detail, okk, maps)
    supervisor_confirmation = build_supervisor_confirmation(may_detail)
    trajectories = build_trajectories(may_detail)
    tables = {
        "Методология": methodology(),
        "Ограничения": limitations(),
        "Кадровые_эпизоды": episodes.drop(columns=["ID сотрудника raw"]),
        "Обучение": learning,
        "KPI_помесячно": kpi,
        "Клиентские_цели_KPI": client_kpi_targets(facts),
        "ОКК_проверки": okk,
        "Полевые_визиты": field_visits,
        "Контроль_RTM": rtm_audit,
        "Покрытие_RTM": rtm_coverage,
        "Действия_руководителя": actions,
        "Подтверждение_СВ": supervisor_confirmation,
        "Мини_замер_времени": time_measurement(),
        "Источники_замера": time_research_sources(),
        "Экономика_допущения": economics_assumptions(episodes),
        "Экономика_формулы": economics_formulas(),
        "Источники_экономики": economics_sources(),
        "Когорты_сводка": cohort_summary,
        "Когорты_сотрудники": cohort_detail,
        "Воронка_6м": cohort_funnel,
        "ОКК_первые14": okk_coverage,
    }
    export_excel(tables)
    print(f"Created: {OUTPUT}")
    print(f"Episodes: {len(episodes)}")
    print(f"Learning rows: {len(learning)}")
    print(f"KPI rows: {len(kpi)}")
    print(f"OKK checks: {len(okk)}")
    print(f"Expanded cohorts: {len(cohort_detail)}")
    print(f"May cohort: {len(may_detail)}")
    print(f"Demo actions: {len(actions)}")


if __name__ == "__main__":
    main()
