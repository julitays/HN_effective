from __future__ import annotations

import json
import re
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from scripts.cache_utils import load_or_build_parquet_cache, source_set_digest
from scripts.staffing_utils import normalize_name


MONTHS = {
    "январ": 1,
    "феврал": 2,
    "март": 3,
    "апрел": 4,
    "май": 5,
    "мая": 5,
    "июн": 6,
    "июл": 7,
    "август": 8,
    "сентябр": 9,
    "октябр": 10,
    "ноябр": 11,
    "декабр": 12,
}
UNMATCHED_DIRECTORY_NAME = "Нет привязки в справочнике"
RTM_CACHE_VERSION = "2"


def _normalize_code(series: pd.Series) -> pd.Series:
    result = series.astype("string").str.strip().str.replace(r"\.0+$", "", regex=True)
    return result.replace({"": pd.NA, "nan": pd.NA, "None": pd.NA, "<NA>": pd.NA})


def _normalize_bool(series: pd.Series) -> pd.Series:
    values = series.astype("string").str.strip().str.lower()
    return values.map(
        {
            "true": True,
            "1": True,
            "да": True,
            "yes": True,
            "false": False,
            "0": False,
            "нет": False,
            "no": False,
        }
    ).astype("boolean")


def _year_month_from_filename(path: Path) -> int:
    stem = path.stem.lower().replace("ё", "е")
    month = next((value for token, value in MONTHS.items() if token in stem), None)
    year_match = re.search(r"(?:20)?(\d{2})(?!\d)", stem)
    if month is None or year_match is None:
        raise ValueError(f"Не удалось определить месяц файла логинов: {path.name}")
    year = 2000 + int(year_match.group(1))
    return year * 100 + month


def _rtm_source_digest(source: Path, rtm_root: Path) -> str:
    return source_set_digest([source], rtm_root, RTM_CACHE_VERSION)


def _parse_rtm_source(source: Path, rtm_root: Path) -> tuple[pd.DataFrame, dict]:
    workbook = load_workbook(source, read_only=True, data_only=True)
    try:
        worksheet = workbook.active
        worksheet.reset_dimensions()
        rows = worksheet.iter_rows(values_only=True)
        first_row = next(rows, None)
        if first_row is None:
            raise ValueError(f"RTM {source.name}: файл пуст")
        filter_text = str(first_row[0] or "")
        next(rows, None)
        header_row = next(rows, None)
        if header_row is None:
            raise ValueError(f"RTM {source.name}: отсутствует строка заголовков")
        headers = list(header_row)
        while headers and headers[-1] is None:
            headers.pop()

        required = [
            "route_date",
            "route_name",
            "employee_id",
            "shop_code",
            "agg_visit_id",
            "visit_date",
            "visit_id",
            "is_complete",
            "is_confirmed",
        ]
        optional = ["region", "business_unit", "sales_group", "city_name"]
        missing = [column for column in required if column not in headers]
        if missing:
            raise ValueError(f"RTM {source.name}: отсутствуют поля {missing}")
        selected = required + [column for column in optional if column in headers]
        positions = [headers.index(column) for column in selected]
        records = ([row[position] for position in positions] for row in rows)
        raw = pd.DataFrame.from_records(records, columns=selected)
    finally:
        workbook.close()

    relative_name = source.relative_to(rtm_root).as_posix()
    frame = pd.DataFrame(index=raw.index)
    frame["Дата визита"] = pd.to_datetime(raw["visit_date"], errors="coerce").dt.normalize()
    frame["Код RTM"] = _normalize_code(raw["employee_id"])
    frame["ТТ"] = _normalize_code(raw["shop_code"])
    frame["Ключ визита RTM"] = _normalize_code(raw["visit_id"]).combine_first(
        _normalize_code(raw["agg_visit_id"])
    )
    frame["Маршрут RTM"] = raw["route_name"].astype("string")
    for source_column, output_column in [
        ("region", "Регион RTM"),
        ("business_unit", "BU RTM"),
        ("sales_group", "SG RTM"),
        ("city_name", "Город RTM"),
    ]:
        frame[output_column] = (
            raw[source_column].astype("string")
            if source_column in raw.columns
            else pd.Series(pd.NA, index=raw.index, dtype="string")
        )
    frame["Визит выполнен"] = _normalize_bool(raw["is_complete"])
    frame["Визит подтверждён"] = _normalize_bool(raw["is_confirmed"])
    frame["Файл RTM"] = relative_name

    observed = frame["Дата визита"].dropna()
    audit = {
        "Файл RTM": relative_name,
        "Фильтр выгрузки": filter_text,
        "Строк": len(frame),
        "Минимальная дата": observed.min() if not observed.empty else pd.NaT,
        "Максимальная дата": observed.max() if not observed.empty else pd.NaT,
        "Возможен лимит выгрузки": len(frame) >= 150_000,
    }
    return frame, audit


def _load_rtm_source(
    source: Path,
    rtm_root: Path,
    cache_root: Path | None,
) -> tuple[pd.DataFrame, dict, Path | None]:
    if cache_root is None:
        frame, audit = _parse_rtm_source(source, rtm_root)
        audit["Режим чтения"] = "Excel"
        return frame, audit, None

    cache_root.mkdir(parents=True, exist_ok=True)
    cache_key = _rtm_source_digest(source, rtm_root)
    cache_path = cache_root / f"{cache_key}.parquet"
    metadata_path = cache_root / f"{cache_key}.json"
    if cache_path.exists() and metadata_path.exists():
        frame = pd.read_parquet(cache_path)
        with metadata_path.open(encoding="utf-8") as handle:
            audit = json.load(handle)
        for column in ["Минимальная дата", "Максимальная дата"]:
            audit[column] = pd.to_datetime(audit.get(column), errors="coerce")
        audit["Режим чтения"] = "Кеш parquet"
        return frame, audit, cache_path

    frame, audit = _parse_rtm_source(source, rtm_root)
    temporary_parquet = cache_path.with_suffix(".tmp.parquet")
    temporary_metadata = metadata_path.with_suffix(".tmp.json")
    try:
        frame.to_parquet(temporary_parquet, index=False, compression="snappy")
        serializable_audit = {
            key: value.isoformat() if isinstance(value, pd.Timestamp) else value
            for key, value in audit.items()
        }
        with temporary_metadata.open("w", encoding="utf-8") as handle:
            json.dump(serializable_audit, handle, ensure_ascii=False, indent=2)
        temporary_parquet.replace(cache_path)
        temporary_metadata.replace(metadata_path)
    except Exception as exc:
        temporary_parquet.unlink(missing_ok=True)
        temporary_metadata.unlink(missing_ok=True)
        print(f"  RTM: кеш для {source.name} не записан — {exc}")
        cache_path = None
    audit["Режим чтения"] = "Excel → кеш"
    return frame, audit, cache_path


def _unique_name_id_map(frame: pd.DataFrame, name_col: str, id_col: str) -> dict[str, str]:
    if frame is None or frame.empty or name_col not in frame.columns or id_col not in frame.columns:
        return {}
    work = frame[[name_col, id_col]].dropna().copy()
    work["ФИО norm"] = work[name_col].map(normalize_name)
    work[id_col] = work[id_col].astype("string").str.strip()
    work = work[work["ФИО norm"].ne("") & work[id_col].ne("")]
    grouped = work.groupby("ФИО norm")[id_col].agg(lambda values: tuple(sorted(set(values.dropna()))))
    return {name: ids[0] for name, ids in grouped.items() if len(ids) == 1}


def _person_resolvers(dim_employees: pd.DataFrame) -> tuple[dict[str, str], dict[str, str]]:
    if dim_employees is None or dim_employees.empty:
        return {}, {}
    if not {"ID сотрудника", "ФИО"}.issubset(dim_employees.columns):
        return {}, {}
    work = dim_employees[["ID сотрудника", "ФИО"]].dropna().copy()
    work["ID сотрудника"] = work["ID сотрудника"].astype("string").str.strip()
    work["full"] = work["ФИО"].map(normalize_name)

    def unique_map(column: str) -> dict[str, str]:
        grouped = work.dropna(subset=[column]).groupby(column)["ID сотрудника"].agg(
            lambda values: tuple(sorted(set(values.dropna().astype(str))))
        )
        return {name: ids[0] for name, ids in grouped.items() if len(ids) == 1}

    canonical = (
        work.drop_duplicates("ID сотрудника", keep="last")
        .set_index("ID сотрудника")["ФИО"]
        .astype("string")
        .to_dict()
    )
    return unique_map("full"), canonical


def _resolve_person(names: pd.Series, dim_employees: pd.DataFrame, prefix: str) -> tuple[pd.Series, pd.Series]:
    full_map, canonical = _person_resolvers(dim_employees)
    cleaned = names.astype("string").str.strip().replace({"": pd.NA, "nan": pd.NA, "None": pd.NA})
    ids = cleaned.map(normalize_name).map(full_map).astype("string")
    canonical_names = ids.map(canonical)
    canonical_names = canonical_names.where(canonical_names.notna(), UNMATCHED_DIRECTORY_NAME)
    canonical_names = canonical_names.where(cleaned.notna(), pd.NA)
    return ids, canonical_names


def _grouped_mode_text(
    frame: pd.DataFrame,
    group_columns: list[str],
    value_column: str,
) -> pd.Series:
    work = frame[group_columns + [value_column]].dropna(subset=[value_column]).copy()
    work[value_column] = work[value_column].astype("string").str.strip()
    work = work[work[value_column].ne("")]
    if work.empty:
        return pd.Series(dtype="string")
    counts = (
        work.groupby(group_columns + [value_column], dropna=False, sort=False)
        .size()
        .reset_index(name="_count")
        .sort_values(
            group_columns + ["_count", value_column],
            ascending=[True] * len(group_columns) + [False, True],
            kind="mergesort",
        )
        .drop_duplicates(group_columns, keep="first")
    )
    return counts.set_index(group_columns)[value_column]


def load_login_employee_map(
    login_root: Path,
    dim_employees: pd.DataFrame,
    hr_registry: pd.DataFrame | None = None,
    cache_root: Path | None = None,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    sources = sorted(login_root.glob("*.xlsx"))
    if not sources:
        raise FileNotFoundError(f"В {login_root} нет пригодных файлов логинов")

    def build_raw_map() -> pd.DataFrame:
        frames = []
        for source in sources:
            year_month = _year_month_from_filename(source)
            with pd.ExcelFile(source) as workbook:
                for sheet in workbook.sheet_names:
                    raw = workbook.parse(sheet, dtype="string")
                    if raw.empty:
                        continue
                    columns = {str(column).strip().lower(): column for column in raw.columns}
                    code_col = columns.get("код")
                    name_col = next(
                        (
                            columns.get(alias)
                            for alias in ("сотрудник", "фио")
                            if columns.get(alias) is not None
                        ),
                        raw.columns[0],
                    )
                    if code_col is None:
                        continue
                    login_col = columns.get("логин") or columns.get("лог")
                    tm_col = columns.get("тм")
                    sv_col = columns.get("св")
                    part = pd.DataFrame(
                        {
                            "YearMonth": year_month,
                            "Код RTM": _normalize_code(raw[code_col]),
                            "ФИО из логинов": raw[name_col].astype("string").str.strip(),
                            "Логин": raw[login_col].astype("string").str.strip() if login_col else pd.NA,
                            "ТМ из логинов": raw[tm_col].astype("string").str.strip() if tm_col else pd.NA,
                            "Код СВ из логинов": raw[sv_col].astype("string").str.strip() if sv_col else pd.NA,
                            "Файл логинов": source.name,
                            "Лист логинов": sheet,
                        }
                    )
                    part["ФИО norm"] = part["ФИО из логинов"].map(normalize_name)
                    part = part[part["Код RTM"].notna() & part["ФИО norm"].ne("")]
                    frames.append(part)
        if not frames:
            raise FileNotFoundError(f"В {login_root} нет пригодных файлов логинов")
        return pd.concat(frames, ignore_index=True).drop_duplicates()

    cache_key = source_set_digest(sources, login_root, "login-map-v1")
    raw_map, cache_hit = load_or_build_parquet_cache(cache_root, cache_key, build_raw_map)
    print(f"  Логины: {'кеш parquet' if cache_hit else 'прочитаны Excel'}")
    group_columns = ["YearMonth", "Код RTM"]
    grouped_index = raw_map.groupby(group_columns, dropna=False, sort=False)["ФИО norm"].unique()
    grouped = grouped_index.map(
        lambda values: tuple(sorted(value for value in values if pd.notna(value)))
    ).to_frame("ФИО norm варианты")
    for column in [
        "ФИО из логинов",
        "Логин",
        "ТМ из логинов",
        "Код СВ из логинов",
        "Файл логинов",
    ]:
        grouped[column] = _grouped_mode_text(raw_map, group_columns, column)
    grouped = grouped.reset_index()
    grouped["Однозначный код"] = grouped["ФИО norm варианты"].map(len).eq(1)
    grouped["ФИО norm"] = grouped["ФИО norm варианты"].map(
        lambda values: values[0] if len(values) == 1 else ""
    )

    dim_map = _unique_name_id_map(dim_employees, "ФИО", "ID сотрудника")
    hr_map = _unique_name_id_map(hr_registry, "Сотрудник", "ID сотрудника")
    grouped["ID USERS"] = grouped["ФИО norm"].map(dim_map)
    grouped["ID HR"] = grouped["ФИО norm"].map(hr_map)
    grouped["Конфликт ID"] = (
        grouped["ID USERS"].notna()
        & grouped["ID HR"].notna()
        & grouped["ID USERS"].ne(grouped["ID HR"])
    )
    grouped["ID сотрудника"] = grouped["ID USERS"].combine_first(grouped["ID HR"])
    grouped.loc[grouped["Конфликт ID"] | ~grouped["Однозначный код"], "ID сотрудника"] = pd.NA
    grouped["Источник ID"] = pd.NA
    grouped.loc[grouped["ID USERS"].notna() & ~grouped["Конфликт ID"], "Источник ID"] = "USERS"
    grouped.loc[
        grouped["ID USERS"].isna() & grouped["ID HR"].notna() & ~grouped["Конфликт ID"],
        "Источник ID",
    ] = "Кадровая история"
    raw_tm_names = grouped["ТМ из логинов"].copy()
    grouped["ID ТМ из логинов"], grouped["ТМ из логинов"] = _resolve_person(
        grouped["ТМ из логинов"], dim_employees, "LOGIN_TM"
    )
    vacancy_tm = raw_tm_names.astype("string").str.contains("вакан", case=False, na=False)
    grouped.loc[vacancy_tm, "ID ТМ из логинов"] = "NO_TM"
    grouped.loc[vacancy_tm, "ТМ из логинов"] = "Вакансия / нет ТМ"

    audit = grouped.copy()
    mapping = grouped[grouped["Однозначный код"]][
        [
            "YearMonth",
            "Код RTM",
            "ID сотрудника",
            "ФИО из логинов",
            "Логин",
            "Источник ID",
            "ID ТМ из логинов",
            "ТМ из логинов",
            "Код СВ из логинов",
        ]
    ].copy()
    return mapping, audit


def load_rtm_employee_visits(
    rtm_root: Path,
    login_map: pd.DataFrame,
    start: pd.Timestamp | None = None,
    end: pd.Timestamp | None = None,
    cache_root: Path | None = None,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    frames = []
    file_audit = []
    active_cache_files: set[Path] = set()
    for source in sorted(rtm_root.rglob("*.xlsx")):
        frame, audit_row, cache_path = _load_rtm_source(source, rtm_root, cache_root)
        if cache_path is not None:
            active_cache_files.add(cache_path)
            active_cache_files.add(cache_path.with_suffix(".json"))
        file_audit.append(audit_row)
        frames.append(frame)
    if not frames:
        raise FileNotFoundError(f"В {rtm_root} нет файлов RTM")

    if cache_root is not None and cache_root.exists():
        for cached in cache_root.iterdir():
            if cached.suffix in {".parquet", ".json"} and cached not in active_cache_files:
                cached.unlink(missing_ok=True)

    rows = pd.concat(frames, ignore_index=True)
    rows["Визит выполнен"] = rows["Визит выполнен"].astype("boolean")
    rows["Визит подтверждён"] = rows["Визит подтверждён"].astype("boolean")

    missing_visit_date = rows["Дата визита"].isna()
    missing_flags = rows["Визит выполнен"].isna() | rows["Визит подтверждён"].isna()
    quality_audit = []
    for label, mask in [
        ("Удалено: нет фактической даты визита", missing_visit_date),
        ("Удалено: пустой флаг выполнения/подтверждения", missing_flags & ~missing_visit_date),
    ]:
        if mask.any():
            quality_audit.append(
                rows.loc[mask]
                .groupby("Файл RTM", dropna=False)
                .size()
                .reset_index(name="Строк")
                .assign(Проверка=label)
            )

    rows = rows[~missing_visit_date & ~missing_flags].copy()
    if start is not None:
        rows = rows[rows["Дата визита"].ge(pd.Timestamp(start))]
    if end is not None:
        rows = rows[rows["Дата визита"].le(pd.Timestamp(end))]
    rows["YearMonth"] = rows["Дата визита"].dt.year * 100 + rows["Дата визита"].dt.month
    missing_visit_key = rows["Ключ визита RTM"].isna()
    if missing_visit_key.any():
        quality_audit.append(
            rows.loc[missing_visit_key]
            .groupby("Файл RTM", dropna=False)
            .size()
            .reset_index(name="Строк")
            .assign(Проверка="Удалено: нет ID визита RTM")
        )
    rows = rows[~missing_visit_key].copy()
    visits = (
        rows.groupby(["YearMonth", "Ключ визита RTM"], dropna=False)
        .agg(
            **{
                "Дата визита": ("Дата визита", "min"),
                "Код RTM": ("Код RTM", "first"),
                "ТТ": ("ТТ", "first"),
                "Маршрут RTM": ("Маршрут RTM", "first"),
                "Регион RTM": ("Регион RTM", "first"),
                "BU RTM": ("BU RTM", "first"),
                "SG RTM": ("SG RTM", "first"),
                "Город RTM": ("Город RTM", "first"),
                "Визит выполнен": ("Визит выполнен", "max"),
                "Визит подтверждён": ("Визит подтверждён", "max"),
                "Файл RTM": ("Файл RTM", "first"),
            }
        )
        .reset_index()
    )
    visits = visits[
        visits["Визит выполнен"].fillna(False) & visits["Визит подтверждён"].fillna(False)
    ].copy()
    visits = visits.merge(login_map, on=["YearMonth", "Код RTM"], how="left")
    visits["MonthStart"] = visits["Дата визита"].dt.to_period("M").dt.to_timestamp()

    month_audit = (
        visits.groupby("YearMonth", dropna=False)
        .agg(
            **{
                "Подтверждённых визитов": ("Ключ визита RTM", "nunique"),
                "Сопоставлено с сотрудником": ("ID сотрудника", lambda values: values.notna().sum()),
                "Уникальных кодов RTM": ("Код RTM", "nunique"),
                "Уникальных сотрудников": ("ID сотрудника", "nunique"),
            }
        )
        .reset_index()
    )
    month_audit["Покрытие сопоставления"] = (
        month_audit["Сопоставлено с сотрудником"] / month_audit["Подтверждённых визитов"]
    )
    audit_parts = [pd.DataFrame(file_audit), month_audit]
    if quality_audit:
        audit_parts.append(pd.concat(quality_audit, ignore_index=True))
    audit = pd.concat(audit_parts, axis=0, ignore_index=True, sort=False)
    return visits, audit
